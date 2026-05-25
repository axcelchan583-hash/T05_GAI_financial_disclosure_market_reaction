#!/usr/bin/env python3
"""Supplier-spillover sample diagnostic inspired by Qian, Peng, and Li (2025).

The reference paper obtains hundreds of observations because the unit is not
only a GenAI announcing firm. It multiplies announcing firms by listed suppliers.

This script checks whether the same expansion is feasible with local A-share
data. It does not claim identification validity.
"""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from run_v3_genai_specificity_car_pilot import A_SHARE_PAT, compute_car, run_regression  # noqa: E402

OUT_DIR = ROOT / "results" / "v3_supplier_spillover_sample_diagnostic"
DOC_OUT = ROOT / "docs" / "选题" / "20_v3_supplier_spillover_sample_diagnostic_20260522.md"

FORMAL_EVENTS = ROOT / "results" / "v3_event_library_expansion_diagnostic" / "formal_all_events_events_before_car.csv"
IRQA_EVENTS = ROOT / "results" / "v3_event_library_expansion_diagnostic" / "irqa_all_events_supplement_events_before_car.csv"

SC_ROOT = Path("/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/上市公司股东")
NETWORK_ZIP = SC_ROOT / "供应链网络关系指标表153711417(仅供北洋大学使用).zip"
TOP_PURCHASE_ZIP = SC_ROOT / "前五大供应商采购信息表153634799(仅供北洋大学使用).zip"
TOP_SALE_ZIP = SC_ROOT / "前五大客户销售信息表152419693(仅供北洋大学使用).zip"


def z6(x: object) -> str:
    if pd.isna(x):
        return ""
    return str(x).replace(".0", "").strip().zfill(6)


def is_ashare(code: object) -> bool:
    return bool(A_SHARE_PAT.match(z6(code)))


def read_xlsx_from_zip(zip_path: Path, xlsx_name: str) -> pd.DataFrame:
    with ZipFile(zip_path) as zf:
        raw = BytesIO(zf.read(xlsx_name))
    wb = load_workbook(raw, read_only=True, data_only=True)
    ws = wb.active
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    data = rows[3:]
    return pd.DataFrame(data, columns=header)


def load_events(path: Path, layer: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype={"sec_code": str})
    df["customer_code"] = df["sec_code"].map(z6)
    df = df[df["customer_code"].map(is_ashare)].copy()
    df["event_date"] = pd.to_datetime(df["announcement_date"], errors="coerce")
    df = df[df["event_date"].notna()].copy()
    df["event_year"] = df["event_date"].dt.year
    df["event_layer"] = layer
    df["customer_event_id"] = (
        df["customer_code"]
        + "_"
        + df["event_date"].dt.strftime("%Y%m%d")
        + "_"
        + df.get("announcement_id", df.index.to_series()).astype(str)
    )
    return df


def load_network_edges() -> pd.DataFrame:
    net = read_xlsx_from_zip(NETWORK_ZIP, "SC_NetworkRelationsIndex.xlsx")
    net = net[net["StateTypeCode"].eq(1)].copy()
    net["year"] = pd.to_datetime(net["EndDate"], errors="coerce").dt.year
    net["Symbol"] = net["Symbol"].map(z6)
    net["PSCSymbol"] = net["PSCSymbol"].map(z6)
    net = net[net["Symbol"].map(is_ashare) & net["PSCSymbol"].map(is_ashare)].copy()

    # Direction 1: row firm is customer, listed PSC is its supplier.
    d1 = net[net["PSCBusinessRelation"].eq(2)].copy()
    d1 = d1.rename(columns={"Symbol": "customer_code", "PSCSymbol": "supplier_code"})
    d1["edge_source"] = "network_customer_lists_supplier"

    # Direction 2: row firm is supplier, listed PSC is its customer.
    d2 = net[net["PSCBusinessRelation"].eq(1)].copy()
    d2 = d2.rename(columns={"Symbol": "supplier_code", "PSCSymbol": "customer_code"})
    d2["edge_source"] = "network_supplier_lists_customer"

    out = pd.concat([d1[["customer_code", "supplier_code", "year", "edge_source"]], d2[["customer_code", "supplier_code", "year", "edge_source"]]], ignore_index=True)
    out = out[out["customer_code"].ne(out["supplier_code"])].drop_duplicates()
    return out


def load_topfive_edges() -> pd.DataFrame:
    pur = read_xlsx_from_zip(TOP_PURCHASE_ZIP, "SC_TopFivePurchaseInfo.xlsx")
    pur = pur[pur["StateTypeCode"].eq(1) & pur["IsListed"].eq("Y")].copy()
    pur["year"] = pd.to_datetime(pur["EndDate"], errors="coerce").dt.year
    pur["customer_code"] = pur["Symbol"].map(z6)
    pur["supplier_code"] = pur["BusinessSymbol"].map(z6)
    pur = pur[pur["customer_code"].map(is_ashare) & pur["supplier_code"].map(is_ashare)].copy()
    pur["edge_source"] = "topfive_purchase_customer_lists_supplier"

    sale = read_xlsx_from_zip(TOP_SALE_ZIP, "SC_TopFiveSaleInfo.xlsx")
    sale = sale[sale["StateTypeCode"].eq(1) & sale["IsListed"].eq("Y")].copy()
    sale["year"] = pd.to_datetime(sale["EndDate"], errors="coerce").dt.year
    sale["supplier_code"] = sale["Symbol"].map(z6)
    sale["customer_code"] = sale["BusinessSymbol"].map(z6)
    sale = sale[sale["customer_code"].map(is_ashare) & sale["supplier_code"].map(is_ashare)].copy()
    sale["edge_source"] = "topfive_sale_supplier_lists_customer"

    cols = ["customer_code", "supplier_code", "year", "edge_source"]
    out = pd.concat([pur[cols], sale[cols]], ignore_index=True)
    out = out[out["customer_code"].ne(out["supplier_code"])].drop_duplicates()
    return out


def expand_events(events: pd.DataFrame, edges: pd.DataFrame, name: str) -> pd.DataFrame:
    if events.empty or edges.empty:
        return pd.DataFrame()
    tmp = events.merge(edges, on="customer_code", how="left")
    tmp = tmp[tmp["supplier_code"].notna()].copy()
    tmp = tmp[tmp["year"].between(tmp["event_year"] - 5, tmp["event_year"] - 1)].copy()
    tmp["edge_age"] = tmp["event_year"] - tmp["year"]
    tmp = tmp.sort_values(["customer_event_id", "supplier_code", "edge_age"])
    tmp = tmp.drop_duplicates(["customer_event_id", "supplier_code"], keep="first").copy()
    tmp["edge_family"] = name
    # compute_car expects sec_code and announcement_date for the stock whose return is measured.
    tmp["announcing_customer_code"] = tmp["customer_code"]
    tmp["sec_code"] = tmp["supplier_code"]
    tmp["announcement_date"] = tmp["event_date"].dt.strftime("%Y-%m-%d")
    return tmp


def table(df: pd.DataFrame, max_rows: int = 25) -> str:
    if df.empty:
        return "无"
    show = df.head(max_rows).copy()
    cols = list(show.columns)
    lines = ["| " + " | ".join(cols) + " |", "| " + " | ".join(["---"] * len(cols)) + " |"]
    for _, row in show.iterrows():
        vals = [str(row.get(c, "")).replace("\n", " ").replace("|", "\\|")[:220] for c in cols]
        lines.append("| " + " | ".join(vals) + " |")
    return "\n".join(lines)


def summarize(name: str, events: pd.DataFrame, expanded: pd.DataFrame, car: pd.DataFrame | None = None) -> dict:
    out = {
        "sample": name,
        "customer_events": events["customer_event_id"].nunique() if not events.empty and "customer_event_id" in events else 0,
        "customer_firms": events["customer_code"].nunique() if not events.empty and "customer_code" in events else 0,
        "supplier_event_obs": len(expanded),
        "supplier_firms": expanded["supplier_code"].nunique() if not expanded.empty else 0,
        "customers_with_listed_suppliers": expanded["customer_code"].nunique() if not expanded.empty else 0,
    }
    if car is not None:
        out.update(
            {
                "car_linked_supplier_event_obs": len(car),
                "car_linked_supplier_firms": car["supplier_code"].nunique() if not car.empty and "supplier_code" in car else 0,
                "mean_abs_car": car["abs_car_m1_p1"].mean() if not car.empty and "abs_car_m1_p1" in car else None,
            }
        )
    return out


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    formal = load_events(FORMAL_EVENTS, "formal_announcement")
    irqa = load_events(IRQA_EVENTS, "irqa_supplement")

    network_edges = load_network_edges()
    topfive_edges = load_topfive_edges()
    union_edges = pd.concat([network_edges, topfive_edges], ignore_index=True).drop_duplicates(
        ["customer_code", "supplier_code", "year"]
    )

    network_edges.to_csv(OUT_DIR / "listed_supplier_edges_network.csv", index=False, encoding="utf-8-sig")
    topfive_edges.to_csv(OUT_DIR / "listed_supplier_edges_topfive.csv", index=False, encoding="utf-8-sig")
    union_edges.to_csv(OUT_DIR / "listed_supplier_edges_union.csv", index=False, encoding="utf-8-sig")

    expansions = {
        "formal_network": expand_events(formal, network_edges, "network"),
        "formal_topfive": expand_events(formal, topfive_edges, "topfive"),
        "formal_union": expand_events(formal, union_edges, "union"),
        "irqa_network": expand_events(irqa, network_edges, "network"),
        "irqa_topfive": expand_events(irqa, topfive_edges, "topfive"),
        "irqa_union": expand_events(irqa, union_edges, "union"),
    }
    for name, df in expansions.items():
        df.to_csv(OUT_DIR / f"{name}_supplier_event_panel_before_car.csv", index=False, encoding="utf-8-sig")

    # Only run CAR for the smaller formal-union sample now. The IRQA layer needs
    # updated 2026 returns before its full CAR run is informative.
    formal_union_car = compute_car(expansions["formal_union"])
    formal_union_car.to_csv(OUT_DIR / "formal_union_supplier_event_panel_with_car.csv", index=False, encoding="utf-8-sig")
    formal_union_reg = run_regression(formal_union_car)
    formal_union_reg.to_csv(OUT_DIR / "formal_union_supplier_abs_car_on_customer_specificity.csv", index=False, encoding="utf-8-sig")

    summary = pd.DataFrame(
        [
            summarize("formal_network", formal, expansions["formal_network"]),
            summarize("formal_topfive", formal, expansions["formal_topfive"]),
            summarize("formal_union", formal, expansions["formal_union"], formal_union_car),
            summarize("irqa_network", irqa, expansions["irqa_network"]),
            summarize("irqa_topfive", irqa, expansions["irqa_topfive"]),
            summarize("irqa_union", irqa, expansions["irqa_union"]),
        ]
    )
    summary.to_csv(OUT_DIR / "supplier_spillover_sample_summary.csv", index=False, encoding="utf-8-sig")

    examples_cols = [
        "announcing_customer_code",
        "supplier_code",
        "announcement_date",
        "announcement_title",
        "edge_source",
        "year",
        "specificity",
        "abs_car_m1_p1",
        "car_m1_p1",
    ]
    formal_examples = formal_union_car[[c for c in examples_cols if c in formal_union_car.columns]].sort_values(
        "abs_car_m1_p1", ascending=False
    ) if not formal_union_car.empty else formal_union_car

    report = f"""# v3 供应商放大样本诊断

日期：2026-05-22

## 1. 为什么参考文献能做到几百个样本

参考文献 `The Impact of Generative AI Announcements on Suppliers` 的样本单位不是单纯的 GenAI 公告，而是 `supplier-announcement observation`。

它的路径是：

```text
14,941 条新闻/通讯社潜在 GenAI announcement
-> 2,084 条北美上市公司 announcement
-> 人工筛出 254 家公司的首次具体 GenAI initiative
-> 匹配上市供应商
-> 最终 117 个 announcement firm、277 个 supplier、515 个 supplier-announcement observations
```

所以它有几百个样本，主要来自两个放大机制：

1. 事件源是 LexisNexis 新闻稿/通讯社，不是交易所正式公告；
2. 回归单位是供应商-公告，而不是公告公司自身。

## 2. 本地 A 股供应商放大诊断

本轮把 GenAI 事件公司视为客户，用 CSMAR 供应链网络关系、前五大客户/供应商表匹配上市供应商。关系窗口暂按参考文献精神取事件年前 5 年。

{table(summary)}

## 3. 正式公告供应商 CAR 试跑

正式公告自身只有 20 个事件，但乘上上市供应商关系后，`formal_union` 可以形成 `{len(expansions["formal_union"])}` 个 supplier-event observations，并且其中 `{len(formal_union_car)}` 个成功连上供应商 CAR。

粗回归如下。注意这只是数据链路，不是正式识别：

{table(formal_union_reg)}

## 4. CAR 最大的供应商-客户事件例子

{table(formal_examples, max_rows=30)}

## 5. 判断

如果要解释“为什么别人能几百个样本”，答案很明确：他们不是把事件局限在交易所公告，也不是只看宣布公司自身反应。

对本项目来说，有三条可行扩样路线：

1. **供应商反应路线**：复制参考文献的样本单位，把 Y 改成供应商 CAR。优点是样本可放大；缺点是这会变成供应链溢出论文，不再是披露具体性论文。
2. **多渠道沟通路线**：把 IR/互动问答纳入主事件源。优点是文本事件能到约 700 个公司-日期；缺点是需要补齐 2026 年日收益，并处理投资者提问诱导。
3. **年报大样本路线**：把年报 GenAI 具体性作为 X，Y 换成年报披露窗口 ABVOL/CAR 或分析师修正。优点是 1300 多家公司；缺点是年报窗口业绩信息污染更重。
"""
    DOC_OUT.parent.mkdir(parents=True, exist_ok=True)
    DOC_OUT.write_text(report, encoding="utf-8")
    print(f"summary={OUT_DIR / 'supplier_spillover_sample_summary.csv'}", flush=True)
    print(f"report={DOC_OUT}", flush=True)


if __name__ == "__main__":
    main()
