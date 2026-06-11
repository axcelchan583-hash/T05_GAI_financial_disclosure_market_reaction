#!/usr/bin/env python3
"""Run v52 supplier-return benchmark for the GenAI announcement sample."""

from __future__ import annotations

import math
import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import run_v23_cninfo_1055_peer_event_study_20260603 as pe  # noqa: E402


RUN_ID = "v54_v52_supplier_benchmark_20260612"
OUT_DIR = ROOT / "results" / RUN_ID
DOC_PATH = ROOT / "docs" / "empirical_runs" / "116_v54_v52_supplier_benchmark_20260612.md"
V53_EVENTS = ROOT / "results" / "v53_v52_llm_empirical_tables_20260612" / "v52_event_samples.csv"
V53_T2 = ROOT / "results" / "v53_v52_llm_empirical_tables_20260612" / "t2_peer_main_effect.csv"

SC_ROOT = Path("/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/上市公司股东")
NETWORK_ZIP = SC_ROOT / "供应链网络关系指标表153711417(仅供北洋大学使用).zip"
TOP_PURCHASE_ZIP = SC_ROOT / "前五大供应商采购信息表153634799(仅供北洋大学使用).zip"
TOP_SALE_ZIP = SC_ROOT / "前五大客户销售信息表152419693(仅供北洋大学使用).zip"

SAMPLES = ["A_all", "A_first_firm", "Dfw_all", "A_Dfw_stack"]
OUTCOMES = pe.OUTCOMES


def z6(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).replace(".0", "").strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits.zfill(6)[-6:] if digits else ""


def is_ashare(code: object) -> bool:
    return bool(z6(code)) and z6(code)[:3] in {"000", "001", "002", "003", "300", "301", "600", "601", "603", "605", "688"}


def read_xlsx_from_zip(zip_path: Path, xlsx_name: str) -> pd.DataFrame:
    with ZipFile(zip_path) as zf:
        raw = BytesIO(zf.read(xlsx_name))
    wb = load_workbook(raw, read_only=True, data_only=True)
    ws = wb.active
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    data = rows[3:]
    return pd.DataFrame(data, columns=header)


def load_network_edges() -> pd.DataFrame:
    net = read_xlsx_from_zip(NETWORK_ZIP, "SC_NetworkRelationsIndex.xlsx")
    net = net[net["StateTypeCode"].eq(1)].copy()
    net["year"] = pd.to_datetime(net["EndDate"], errors="coerce").dt.year
    net["Symbol"] = net["Symbol"].map(z6)
    net["PSCSymbol"] = net["PSCSymbol"].map(z6)
    net = net[net["Symbol"].map(is_ashare) & net["PSCSymbol"].map(is_ashare)].copy()

    d1 = net[net["PSCBusinessRelation"].eq(2)].copy()
    d1 = d1.rename(columns={"Symbol": "customer_code", "PSCSymbol": "supplier_code"})
    d1["edge_source"] = "network_customer_lists_supplier"

    d2 = net[net["PSCBusinessRelation"].eq(1)].copy()
    d2 = d2.rename(columns={"Symbol": "supplier_code", "PSCSymbol": "customer_code"})
    d2["edge_source"] = "network_supplier_lists_customer"

    cols = ["customer_code", "supplier_code", "year", "edge_source"]
    out = pd.concat([d1[cols], d2[cols]], ignore_index=True)
    out = out[out["customer_code"].ne(out["supplier_code"])].drop_duplicates()
    return out


def load_topfive_edges() -> pd.DataFrame:
    pur = read_xlsx_from_zip(TOP_PURCHASE_ZIP, "SC_TopFivePurchaseInfo.xlsx")
    pur = pur[pur["StateTypeCode"].eq(1) & pur["IsListed"].eq("Y")].copy()
    pur["year"] = pd.to_datetime(pur["EndDate"], errors="coerce").dt.year
    pur["customer_code"] = pur["Symbol"].map(z6)
    pur["supplier_code"] = pur["BusinessSymbol"].map(z6)
    pur = pur[pur["customer_code"].map(is_ashare) & pur["supplier_code"].map(is_ashare)].copy()
    pur["edge_source"] = "topfive_purchase_customer_lists_supplier"

    sale = read_xlsx_from_zip(TOP_SALE_ZIP, "SC_TopFiveSaleInfo.xlsx")
    sale = sale[sale["StateTypeCode"].eq(1) & sale["IsListed"].eq("Y")].copy()
    sale["year"] = pd.to_datetime(sale["EndDate"], errors="coerce").dt.year
    sale["supplier_code"] = sale["Symbol"].map(z6)
    sale["customer_code"] = sale["BusinessSymbol"].map(z6)
    sale = sale[sale["customer_code"].map(is_ashare) & sale["supplier_code"].map(is_ashare)].copy()
    sale["edge_source"] = "topfive_sale_supplier_lists_customer"

    cols = ["customer_code", "supplier_code", "year", "edge_source"]
    out = pd.concat([pur[cols], sale[cols]], ignore_index=True)
    out = out[out["customer_code"].ne(out["supplier_code"])].drop_duplicates()
    return out


def load_events() -> pd.DataFrame:
    events = pd.read_csv(V53_EVENTS, dtype={"focal_code": str, "event_id": str}, low_memory=False)
    events = events[events["sample_name"].isin(SAMPLES)].copy()
    events["focal_code"] = events["focal_code"].map(z6)
    events["customer_code"] = events["focal_code"]
    events["event_date"] = pd.to_datetime(events["event_date"], errors="coerce")
    events["event_year"] = events["event_date"].dt.year
    events = events[events["customer_code"].map(is_ashare) & events["event_date"].notna()].copy()
    return events.drop_duplicates(["sample_name", "event_key"]).copy()


def expand_events(events: pd.DataFrame, edges: pd.DataFrame, edge_family: str) -> pd.DataFrame:
    if events.empty or edges.empty:
        return pd.DataFrame()
    tmp = events.merge(edges, on="customer_code", how="left")
    tmp = tmp[tmp["supplier_code"].notna()].copy()
    tmp = tmp[tmp["year"].between(tmp["event_year"] - 5, tmp["event_year"] - 1)].copy()
    tmp["edge_age"] = tmp["event_year"] - tmp["year"]
    tmp = tmp.sort_values(["sample_name", "event_key", "supplier_code", "edge_age", "edge_source"])
    tmp = tmp.drop_duplicates(["sample_name", "event_key", "supplier_code"], keep="first").copy()
    tmp["edge_family"] = edge_family
    tmp["peer_code"] = tmp["supplier_code"]
    tmp["supplier_code"] = tmp["supplier_code"].map(z6)
    tmp["supplier_event_key"] = tmp["sample_name"] + "::" + tmp["event_key"] + "::" + tmp["supplier_code"]
    return tmp


def build_supplier_panel(events: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    network = load_network_edges()
    topfive = load_topfive_edges()
    union = pd.concat([network, topfive], ignore_index=True).drop_duplicates(["customer_code", "supplier_code", "year"])

    network.to_csv(OUT_DIR / "listed_supplier_edges_network.csv", index=False)
    topfive.to_csv(OUT_DIR / "listed_supplier_edges_topfive.csv", index=False)
    union.to_csv(OUT_DIR / "listed_supplier_edges_union.csv", index=False)

    panels = [
        expand_events(events, network, "network"),
        expand_events(events, topfive, "topfive"),
        expand_events(events, union, "union"),
    ]
    panel = pd.concat([p for p in panels if not p.empty], ignore_index=True)
    coverage = (
        panel.groupby(["sample_name", "event_type", "edge_family"], dropna=False)
        .agg(
            events_with_suppliers=("event_key", "nunique"),
            supplier_event_obs=("supplier_code", "size"),
            supplier_firms=("supplier_code", "nunique"),
            customer_firms_with_suppliers=("customer_code", "nunique"),
        )
        .reset_index()
    )
    sample_inputs = events.groupby(["sample_name", "event_type"])["event_key"].nunique().rename("input_events").reset_index()
    coverage = coverage.merge(sample_inputs, on=["sample_name", "event_type"], how="left")
    coverage["input_events"] = coverage["input_events"].astype(int)
    coverage["event_link_rate"] = coverage["events_with_suppliers"] / coverage["input_events"]
    return panel, coverage


def add_supplier_returns(panel: pd.DataFrame) -> pd.DataFrame:
    stock = pe.load_stock_model()
    out = pe.attach_event_trading_dates(panel, stock)
    out = pe.build_return_measures(out, stock)
    return out


def mean_event_study(panel: pd.DataFrame) -> pd.DataFrame:
    d0 = panel[panel["complete_clean_m1_p1"].eq(1)].copy()
    rows = []
    group_cols = ["sample_name", "event_type", "edge_family"]
    for key, d in d0.groupby(group_cols, dropna=False):
        base = dict(zip(group_cols, key))
        for outcome, label in OUTCOMES:
            rows.append({**base, "outcome": outcome, "outcome_label": label, **pe.clustered_mean(d, outcome)})
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["sample_name", "event_type", "edge_family", "outcome"])
    return out


def compare_competitor_supplier(t9: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if V53_T2.exists():
        t2 = pd.read_csv(V53_T2)
        comp = t2[
            t2["sample_name"].eq("A_all")
            & t2["method_variant"].eq("liu_product_tfidf_same_industry_d_top10")
            & t2["outcome"].eq("peer_car_0_p1_mm")
        ]
        if not comp.empty:
            r = comp.iloc[0]
            rows.append(
                {
                    "side": "product_market_competitors",
                    "sample_name": "A_all",
                    "edge_family": "liu_product_tfidf_same_industry_d_top10",
                    "outcome_label": "CAR[0,+1]",
                    "estimate": r["estimate"],
                    "se": r["se"],
                    "p": r["p"],
                    "events": r["events"],
                    "firms": r["peer_firms"],
                }
            )
    supp = t9[
        t9["sample_name"].eq("A_all")
        & t9["event_type"].eq("A")
        & t9["edge_family"].eq("union")
        & t9["outcome"].eq("peer_car_0_p1_mm")
    ]
    if not supp.empty:
        r = supp.iloc[0]
        rows.append(
            {
                "side": "listed_suppliers",
                "sample_name": "A_all",
                "edge_family": "union",
                "outcome_label": "CAR[0,+1]",
                "estimate": r["estimate"],
                "se": r["se"],
                "p": r["p"],
                "events": r["events"],
                "firms": r["peer_firms"],
            }
        )
    return pd.DataFrame(rows)


def md_table(df: pd.DataFrame, cols: list[str] | None = None, limit: int = 20) -> str:
    if df.empty:
        return "_No rows._"
    out = df.copy()
    if cols is not None:
        out = out[cols]
    out = out.head(limit).copy()
    for col in out.select_dtypes(include="number").columns:
        out[col] = out[col].map(lambda x: round(float(x), 6) if pd.notna(x) else x)
    lines = ["| " + " | ".join(out.columns) + " |", "|" + "|".join("---" for _ in out.columns) + "|"]
    for _, row in out.iterrows():
        lines.append("| " + " | ".join("" if pd.isna(row[c]) else str(row[c]) for c in out.columns) + " |")
    return "\n".join(lines)


def write_doc(coverage: pd.DataFrame, t9: pd.DataFrame, contrast: pd.DataFrame) -> None:
    DOC_PATH.parent.mkdir(parents=True, exist_ok=True)
    main = t9[
        t9["edge_family"].eq("union")
        & t9["outcome"].isin(["peer_ar0_mm", "peer_car_0_p1_mm"])
    ].sort_values(["sample_name", "event_type", "outcome"])
    lines = [
        "# v54 v52 supplier benchmark",
        "",
        "## Scope",
        "",
        "- Input events: v52/v3.3 LLM-coded sample produced in v53, after valid A-share code and event-date filters.",
        "- Supplier links: CSMAR supply-chain network plus top-five supplier/customer tables, event-year minus 1 to minus 5.",
        "- Return measure: same market-model abnormal returns as the competitor event study.",
        "- This is a benchmark, not the headline identification table; listed supplier coverage is sparse in A-share data.",
        "",
        "## Supplier Coverage",
        "",
        md_table(
            coverage.sort_values(["sample_name", "event_type", "edge_family"]),
            [
                "sample_name",
                "event_type",
                "edge_family",
                "input_events",
                "events_with_suppliers",
                "event_link_rate",
                "supplier_event_obs",
                "supplier_firms",
                "customer_firms_with_suppliers",
            ],
            40,
        ),
        "",
        "## T9 Supplier Event Study",
        "",
        md_table(
            main,
            ["sample_name", "event_type", "edge_family", "outcome_label", "estimate", "se", "p", "nobs", "events", "peer_firms", "median", "positive_share"],
            40,
        ),
        "",
        "## Competitor vs Supplier Sign Check",
        "",
        md_table(contrast, ["side", "sample_name", "edge_family", "outcome_label", "estimate", "se", "p", "events", "firms"], 10),
        "",
        "## Output Files",
        "",
        f"- `{OUT_DIR.relative_to(ROOT)}/supplier_coverage_summary.csv`",
        f"- `{OUT_DIR.relative_to(ROOT)}/supplier_event_panel.csv.gz`",
        f"- `{OUT_DIR.relative_to(ROOT)}/supplier_event_panel_with_returns.csv.gz`",
        f"- `{OUT_DIR.relative_to(ROOT)}/t9_supplier_event_study.csv`",
        f"- `{OUT_DIR.relative_to(ROOT)}/competitor_vs_supplier_sign_check.csv`",
        f"- `{OUT_DIR.relative_to(ROOT)}/{RUN_ID}.xlsx`",
    ]
    DOC_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    events = load_events()
    panel, coverage = build_supplier_panel(events)
    panel.to_csv(OUT_DIR / "supplier_event_panel.csv.gz", index=False)
    coverage.to_csv(OUT_DIR / "supplier_coverage_summary.csv", index=False)

    with_returns = add_supplier_returns(panel)
    with_returns.to_csv(OUT_DIR / "supplier_event_panel_with_returns.csv.gz", index=False)

    t9 = mean_event_study(with_returns)
    contrast = compare_competitor_supplier(t9)
    t9.to_csv(OUT_DIR / "t9_supplier_event_study.csv", index=False)
    contrast.to_csv(OUT_DIR / "competitor_vs_supplier_sign_check.csv", index=False)

    xlsx = OUT_DIR / f"{RUN_ID}.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        coverage.to_excel(writer, sheet_name="Coverage", index=False)
        t9.to_excel(writer, sheet_name="T9_supplier", index=False)
        contrast.to_excel(writer, sheet_name="Sign_check", index=False)

    write_doc(coverage, t9, contrast)

    main_rows = t9[
        t9["sample_name"].eq("A_all")
        & t9["edge_family"].eq("union")
        & t9["outcome"].isin(["peer_ar0_mm", "peer_car_0_p1_mm"])
    ]
    print(f"wrote {OUT_DIR}", flush=True)
    print(f"wrote {DOC_PATH}", flush=True)
    print("\nA_all supplier benchmark:")
    print(main_rows.to_string(index=False), flush=True)
    print("\nCompetitor vs supplier:")
    print(contrast.to_string(index=False), flush=True)


if __name__ == "__main__":
    main()
