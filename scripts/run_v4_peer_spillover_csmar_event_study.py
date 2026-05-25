#!/usr/bin/env python3
"""Run the v4 peer-spillover event study with CSMAR daily returns.

This replaces the earlier Sina K-line pilot with local CSMAR daily stock and
index returns. It reads only the peer stocks needed by the v4 X sample, caches
the filtered return panel, computes AR[0], CAR[0,+1], and CAR[-1,+1], and runs
the same event fixed-effect main-effect regressions.
"""

from __future__ import annotations

import argparse
import math
import os
import re
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
X_PATH = ROOT / "results" / "v4_peer_spillover_x_pilot" / "v4_peer_event_x_same_industry_d_top10.csv"
OUT_DIR = ROOT / "results" / "v4_peer_spillover_csmar_event_study"

DAILY_DIR = Path(
    "/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/"
    "上市公司财务信息/_daily_2021_2026"
)
DAILY_DTA_PATCH_DIR = Path(
    "/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/"
    "上市公司财务信息/_daily_2026_02_14_2026_05_22_patch_dta"
)
IDX_ROOT = Path(
    "/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/"
    "上市公司财务信息"
)
INDEX_DTA_PATCH_DIR = Path(
    "/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/"
    "上市公司财务信息/_idx_2026_02_14_2026_05_22_patch_dta"
)

MAX_PEER_RANK = int(os.environ.get("V4_MAX_PEER_RANK", "5"))
MIN_ESTIMATION_OBS = int(os.environ.get("V4_MIN_ESTIMATION_OBS", "120"))


def code6(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"\D", "", text)
    return text.zfill(6) if text else ""


def to_float(value: object) -> float:
    if value is None:
        return math.nan
    try:
        return float(value)
    except (TypeError, ValueError):
        return math.nan


def to_int(value: object) -> int | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def read_x_sample(max_peer_rank: int) -> pd.DataFrame:
    x = pd.read_csv(X_PATH, dtype={"focal_code": str, "peer_code": str})
    x["focal_code"] = x["focal_code"].map(code6)
    x["peer_code"] = x["peer_code"].map(code6)
    x["raw_event_date"] = pd.to_datetime(x["event_date"], errors="coerce")
    x["peer_rank"] = pd.to_numeric(x["peer_rank"], errors="coerce")
    for col in ["specificity", "product_similarity", "x_specificity_similarity"]:
        x[col] = pd.to_numeric(x[col], errors="coerce")
    x = x[
        x["peer_code"].ne("")
        & x["raw_event_date"].notna()
        & x["peer_rank"].le(max_peer_rank)
    ].copy()
    x["event_id"] = x["event_id"].astype(str)
    return x


def xlsx_paths_daily() -> list[Path]:
    return sorted(
        p
        for p in DAILY_DIR.glob("TRD_Dalyr*.xlsx")
        if "[DES]" not in p.name
    )


def dta_paths_daily() -> list[Path]:
    return sorted(
        p
        for p in DAILY_DTA_PATCH_DIR.glob("TRD_Dalyr*.dta")
        if "[DES]" not in p.name
    )


def xlsx_paths_index() -> list[Path]:
    return sorted(
        p
        for p in IDX_ROOT.glob("_idx_batch*/IDX_Idxtrd*.xlsx")
        if "[DES]" not in p.name
    )


def dta_path_index() -> Path:
    return INDEX_DTA_PATCH_DIR / "TRD_Index.dta"


def iter_rows_from_xlsx(path: Path, max_col: int):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # Some CSMAR xlsx files have a broken worksheet dimension cache in read-only
    # mode, which makes openpyxl expose only column A unless max_col is explicit.
    ws.reset_dimensions()
    rows = ws.iter_rows(values_only=True, max_col=max_col)
    header = next(rows)
    colmap = {str(name): idx for idx, name in enumerate(header)}
    try:
        next(rows)  # Chinese label row
        next(rows)  # unit row
        for row in rows:
            yield colmap, row
    finally:
        wb.close()


def build_stock_return_cache_from_dta(target_codes: set[str]) -> pd.DataFrame:
    chunks: list[pd.DataFrame] = []
    cols = [
        "Stkcd",
        "Trddt",
        "Dretwd",
        "Dretnd",
        "ChangeRatio",
        "Markettype",
        "Trdsta",
        "LimitStatus",
    ]
    for path in dta_paths_daily():
        print(f"Reading CSMAR stock daily dta file: {path.name}", flush=True)
        for chunk in pd.read_stata(path, columns=cols, chunksize=200_000):
            chunk = chunk.copy()
            chunk["peer_code"] = chunk["Stkcd"].map(code6)
            chunk = chunk[chunk["peer_code"].isin(target_codes)].copy()
            if chunk.empty:
                continue
            out = pd.DataFrame(
                {
                    "peer_code": chunk["peer_code"],
                    "date": pd.to_datetime(chunk["Trddt"], errors="coerce"),
                    "ret_wd": pd.to_numeric(chunk["Dretwd"], errors="coerce"),
                    "ret_nd": pd.to_numeric(chunk["Dretnd"], errors="coerce"),
                    "change_ratio": pd.to_numeric(chunk["ChangeRatio"], errors="coerce"),
                    "market_type": pd.to_numeric(chunk["Markettype"], errors="coerce"),
                    "trdsta": pd.to_numeric(chunk["Trdsta"], errors="coerce"),
                    "limit_status": pd.to_numeric(chunk["LimitStatus"], errors="coerce"),
                }
            )
            chunks.append(out)
        total = sum(len(chunk) for chunk in chunks)
        print(f"  cached rows so far: {total:,}", flush=True)

    if not chunks:
        return pd.DataFrame(
            columns=[
                "peer_code",
                "date",
                "ret_wd",
                "ret_nd",
                "change_ratio",
                "market_type",
                "trdsta",
                "limit_status",
            ]
        )
    return pd.concat(chunks, ignore_index=True)


def build_stock_return_cache_from_xlsx(target_codes: set[str]) -> pd.DataFrame:
    keep_cols = {
        "Stkcd",
        "Trddt",
        "Dretwd",
        "Dretnd",
        "ChangeRatio",
        "Markettype",
        "Trdsta",
        "LimitStatus",
    }
    rows: list[dict[str, object]] = []
    for path in xlsx_paths_daily():
        print(f"Reading CSMAR stock daily xlsx file: {path.name}", flush=True)
        for colmap, row in iter_rows_from_xlsx(path, max_col=24):
            code = code6(row[colmap["Stkcd"]])
            if code not in target_codes:
                continue
            rec = {col: row[colmap[col]] if col in colmap else None for col in keep_cols}
            date = pd.to_datetime(rec["Trddt"], errors="coerce")
            if pd.isna(date):
                continue
            rows.append(
                {
                    "peer_code": code,
                    "date": date.strftime("%Y-%m-%d"),
                    "ret_wd": to_float(rec["Dretwd"]),
                    "ret_nd": to_float(rec["Dretnd"]),
                    "change_ratio": to_float(rec["ChangeRatio"]),
                    "market_type": to_int(rec["Markettype"]),
                    "trdsta": to_int(rec["Trdsta"]),
                    "limit_status": to_int(rec["LimitStatus"]),
                }
            )
        print(f"  cached rows so far: {len(rows):,}", flush=True)
    return pd.DataFrame(rows)


def build_stock_return_cache(target_codes: set[str], refresh: bool = False) -> pd.DataFrame:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = OUT_DIR / "v4_csmar_peer_daily_returns_subset.csv"
    if cache_path.exists() and not refresh:
        out = pd.read_csv(cache_path, dtype={"peer_code": str})
        out["date"] = pd.to_datetime(out["date"])
        return out

    if dta_paths_daily():
        out = build_stock_return_cache_from_dta(target_codes)
    else:
        out = build_stock_return_cache_from_xlsx(target_codes)
    if out.empty:
        out = pd.DataFrame(
            columns=[
                "peer_code",
                "date",
                "ret_wd",
                "ret_nd",
                "change_ratio",
                "market_type",
                "trdsta",
                "limit_status",
            ]
        )
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out = out.dropna(subset=["date", "ret_wd"]).drop_duplicates(["peer_code", "date"])
    out = out.sort_values(["peer_code", "date"]).reset_index(drop=True)
    out.to_csv(cache_path, index=False)
    return out


def build_market_return_cache_from_dta(index_code: str) -> pd.DataFrame:
    path = dta_path_index()
    print(f"Reading CSMAR index dta file: {path.name}", flush=True)
    raw = pd.read_stata(path, columns=["Indexcd", "Trddt", "Retindex"])
    raw["index_code"] = raw["Indexcd"].map(code6)
    raw = raw[raw["index_code"].eq(index_code)].copy()
    return pd.DataFrame(
        {
            "index_code": raw["index_code"],
            "date": pd.to_datetime(raw["Trddt"], errors="coerce"),
            # TRD_Index.dta Retindex is already in decimal-return units.
            "mkt_ret": pd.to_numeric(raw["Retindex"], errors="coerce"),
            "index_name": "",
        }
    )


def build_market_return_cache_from_xlsx(index_code: str) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for path in xlsx_paths_index():
        print(f"Reading CSMAR index xlsx file: {path.parent.name}/{path.name}", flush=True)
        for colmap, row in iter_rows_from_xlsx(path, max_col=10):
            code = code6(row[colmap["Indexcd"]])
            if code != index_code:
                continue
            date = pd.to_datetime(row[colmap["Idxtrd01"]], errors="coerce")
            if pd.isna(date):
                continue
            rows.append(
                {
                    "index_code": code,
                    "date": date.strftime("%Y-%m-%d"),
                    "mkt_ret": to_float(row[colmap["Idxtrd08"]]) / 100.0,
                    "index_name": row[colmap["Idxtrd09"]],
                }
            )
    return pd.DataFrame(rows)


def build_market_return_cache(refresh: bool = False, index_code: str = "000001") -> pd.DataFrame:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = OUT_DIR / f"v4_csmar_market_returns_{index_code}.csv"
    if cache_path.exists() and not refresh:
        out = pd.read_csv(cache_path)
        out["date"] = pd.to_datetime(out["date"])
        return out

    if dta_path_index().exists():
        out = build_market_return_cache_from_dta(index_code)
    else:
        out = build_market_return_cache_from_xlsx(index_code)
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out = out.dropna(subset=["date", "mkt_ret"]).drop_duplicates(["index_code", "date"])
    out = out.sort_values("date").reset_index(drop=True)
    out.to_csv(cache_path, index=False)
    return out


def choose_event_trading_date(
    raw_date: pd.Timestamp,
    after_close_or_weekend: bool,
    trading_dates: list[pd.Timestamp],
) -> pd.Timestamp | None:
    if pd.isna(raw_date):
        return None
    raw_date = pd.Timestamp(raw_date).normalize()
    if after_close_or_weekend:
        candidates = [d for d in trading_dates if d > raw_date]
    else:
        candidates = [d for d in trading_dates if d >= raw_date]
    return candidates[0] if candidates else None


def window_dates(
    trading_dates: list[pd.Timestamp],
    date_index: dict[pd.Timestamp, int],
    event_date: pd.Timestamp,
    start: int,
    end: int,
) -> list[pd.Timestamp]:
    idx = date_index[event_date]
    lo = max(0, idx + start)
    hi = min(len(trading_dates), idx + end + 1)
    return trading_dates[lo:hi]


def compute_event_returns(x: pd.DataFrame, stock: pd.DataFrame, mkt: pd.DataFrame) -> pd.DataFrame:
    stock = stock.copy()
    stock["date"] = pd.to_datetime(stock["date"])
    mkt = mkt.copy()
    mkt["date"] = pd.to_datetime(mkt["date"])
    ret = stock.merge(mkt[["date", "mkt_ret"]], on="date", how="inner")
    ret = ret.dropna(subset=["ret_wd", "mkt_ret"])

    trading_dates = sorted(mkt["date"].dropna().unique())
    trading_dates = [pd.Timestamp(d).normalize() for d in trading_dates]
    date_index = {d: i for i, d in enumerate(trading_dates)}
    ret_by_peer = {
        code: sub.sort_values("date").copy()
        for code, sub in ret.groupby("peer_code", sort=False)
    }

    rows: list[dict[str, object]] = []
    for _, ev in x.iterrows():
        peer = ev["peer_code"]
        firm_ret = ret_by_peer.get(peer)
        if firm_ret is None or firm_ret.empty:
            continue

        event_date = choose_event_trading_date(
            ev["raw_event_date"],
            bool(ev.get("any_after_close", False) or ev.get("any_weekend_answer", False)),
            trading_dates,
        )
        if event_date is None or event_date not in date_index:
            continue

        est_dates = set(window_dates(trading_dates, date_index, event_date, -210, -11))
        w_m1_p1 = set(window_dates(trading_dates, date_index, event_date, -1, 1))
        w_0_p1 = set(window_dates(trading_dates, date_index, event_date, 0, 1))
        w_0 = {event_date}

        est = firm_ret[firm_ret["date"].isin(est_dates)].copy()
        if len(est) < MIN_ESTIMATION_OBS:
            continue

        xmat = np.column_stack([np.ones(len(est)), est["mkt_ret"].to_numpy()])
        y = est["ret_wd"].to_numpy()
        alpha, beta = np.linalg.lstsq(xmat, y, rcond=None)[0]

        event_window = firm_ret[firm_ret["date"].isin(w_m1_p1)].copy()
        if event_window.empty:
            continue
        event_window["abret"] = event_window["ret_wd"] - (
            alpha + beta * event_window["mkt_ret"]
        )

        def sum_col(dates: set[pd.Timestamp], col: str) -> float:
            sub = event_window[event_window["date"].isin(dates)]
            return float(sub[col].sum()) if not sub.empty else math.nan

        def count_dates(dates: set[pd.Timestamp]) -> int:
            return int(event_window["date"].isin(dates).sum())

        def normal_trading_flag(dates: set[pd.Timestamp]) -> int:
            sub = event_window[event_window["date"].isin(dates)]
            if sub.empty:
                return 0
            return int((sub["trdsta"].fillna(-999).astype(int) == 1).all())

        def no_limit_flag(dates: set[pd.Timestamp]) -> int:
            sub = event_window[event_window["date"].isin(dates)]
            if sub.empty:
                return 0
            return int((sub["limit_status"].fillna(0).astype(int) == 0).all())

        rec = ev.to_dict()
        rec.update(
            {
                "event_date_for_car": event_date.strftime("%Y-%m-%d"),
                "estimation_obs": int(len(est)),
                "peer_alpha": float(alpha),
                "peer_beta": float(beta),
                "peer_ar0": sum_col(w_0, "abret"),
                "peer_car_0_p1": sum_col(w_0_p1, "abret"),
                "peer_car_m1_p1": sum_col(w_m1_p1, "abret"),
                "peer_abs_ar0": abs(sum_col(w_0, "abret")),
                "peer_abs_car_0_p1": abs(sum_col(w_0_p1, "abret")),
                "peer_abs_car_m1_p1": abs(sum_col(w_m1_p1, "abret")),
                "peer_raw_ret_0": sum_col(w_0, "ret_wd"),
                "peer_raw_ret_0_p1": sum_col(w_0_p1, "ret_wd"),
                "peer_raw_ret_m1_p1": sum_col(w_m1_p1, "ret_wd"),
                "mkt_ret_0": sum_col(w_0, "mkt_ret"),
                "mkt_ret_0_p1": sum_col(w_0_p1, "mkt_ret"),
                "mkt_ret_m1_p1": sum_col(w_m1_p1, "mkt_ret"),
                "obs_ar0": count_dates(w_0),
                "obs_car_0_p1": count_dates(w_0_p1),
                "obs_car_m1_p1": count_dates(w_m1_p1),
                "normal_trading_ar0": normal_trading_flag(w_0),
                "normal_trading_0_p1": normal_trading_flag(w_0_p1),
                "normal_trading_m1_p1": normal_trading_flag(w_m1_p1),
                "no_limit_ar0": no_limit_flag(w_0),
                "no_limit_0_p1": no_limit_flag(w_0_p1),
                "no_limit_m1_p1": no_limit_flag(w_m1_p1),
            }
        )
        rows.append(rec)

    return pd.DataFrame(rows)


def regression_row(label: str, outcome: str, df: pd.DataFrame) -> dict[str, object]:
    formula = f"{outcome} ~ product_similarity + x_specificity_similarity + C(event_id)"
    data = df.dropna(
        subset=[outcome, "product_similarity", "x_specificity_similarity", "event_id"]
    ).copy()
    model = smf.ols(formula, data=data).fit(
        cov_type="cluster",
        cov_kwds={"groups": data["event_id"]},
    )
    term = "x_specificity_similarity"
    return {
        "model": label,
        "outcome": outcome,
        "formula": formula,
        "coef_x": model.params.get(term, np.nan),
        "se_x_cluster_event": model.bse.get(term, np.nan),
        "t_x": model.tvalues.get(term, np.nan),
        "p_x": model.pvalues.get(term, np.nan),
        "nobs": int(model.nobs),
        "events": int(data["event_id"].nunique()),
        "peer_firms": int(data["peer_code"].nunique()),
        "r2": float(model.rsquared),
    }


def run_regressions(panel: pd.DataFrame) -> pd.DataFrame:
    if panel.empty:
        return pd.DataFrame(
            [
                {
                    "model": "not_run",
                    "outcome": "",
                    "error": "empty event-return panel; check return-data date coverage against event dates",
                    "nobs": 0,
                    "events": 0,
                    "peer_firms": 0,
                }
            ]
        )

    rows: list[dict[str, object]] = []
    specs = [
        ("raw_ar0", "peer_ar0", panel),
        ("raw_car_0_p1", "peer_car_0_p1", panel),
        ("raw_car_m1_p1", "peer_car_m1_p1", panel),
    ]

    clean = panel[
        (panel["normal_trading_m1_p1"] == 1)
        & (panel["no_limit_m1_p1"] == 1)
        & (panel["obs_car_m1_p1"] == 3)
    ].copy()
    specs.extend(
        [
            ("clean_ar0", "peer_ar0", clean),
            ("clean_car_0_p1", "peer_car_0_p1", clean),
            ("clean_car_m1_p1", "peer_car_m1_p1", clean),
        ]
    )

    for label, outcome, data in specs:
        if data.empty:
            continue
        try:
            rows.append(regression_row(label, outcome, data))
        except Exception as exc:  # noqa: BLE001
            rows.append(
                {
                    "model": label,
                    "outcome": outcome,
                    "error": str(exc),
                    "nobs": len(data),
                    "events": data["event_id"].nunique() if "event_id" in data else np.nan,
                    "peer_firms": data["peer_code"].nunique() if "peer_code" in data else np.nan,
                }
            )
    return pd.DataFrame(rows)


def write_summary(
    x: pd.DataFrame,
    stock: pd.DataFrame,
    mkt: pd.DataFrame,
    panel: pd.DataFrame,
    suffix: str = "",
) -> None:
    stock_target = stock[stock["peer_code"].isin(set(x["peer_code"].dropna().unique()))].copy()
    event_min = x["raw_event_date"].min()
    event_max = x["raw_event_date"].max()
    stock_min = stock_target["date"].min() if not stock_target.empty else pd.NaT
    stock_max = stock_target["date"].max() if not stock_target.empty else pd.NaT
    mkt_min = mkt["date"].min() if not mkt.empty else pd.NaT
    mkt_max = mkt["date"].max() if not mkt.empty else pd.NaT

    summary = pd.DataFrame(
        [
            {"metric": "x_peer_event_obs", "value": len(x)},
            {"metric": "x_events", "value": x["event_id"].nunique()},
            {"metric": "x_focal_firms", "value": x["focal_code"].nunique()},
            {"metric": "x_peer_firms", "value": x["peer_code"].nunique()},
            {"metric": "event_min_date", "value": str(event_min.date()) if pd.notna(event_min) else ""},
            {"metric": "event_max_date", "value": str(event_max.date()) if pd.notna(event_max) else ""},
            {"metric": "stock_return_rows", "value": len(stock_target)},
            {"metric": "stock_return_peer_firms", "value": stock_target["peer_code"].nunique()},
            {"metric": "stock_return_min_date", "value": str(stock_min.date()) if pd.notna(stock_min) else ""},
            {"metric": "stock_return_max_date", "value": str(stock_max.date()) if pd.notna(stock_max) else ""},
            {"metric": "market_return_days", "value": len(mkt)},
            {"metric": "market_return_min_date", "value": str(mkt_min.date()) if pd.notna(mkt_min) else ""},
            {"metric": "market_return_max_date", "value": str(mkt_max.date()) if pd.notna(mkt_max) else ""},
            {"metric": "event_return_obs", "value": len(panel)},
            {"metric": "event_return_events", "value": panel["event_id"].nunique() if not panel.empty else 0},
            {"metric": "event_return_peer_firms", "value": panel["peer_code"].nunique() if not panel.empty else 0},
            {
                "metric": "clean_m1_p1_obs",
                "value": int(
                    (
                        (panel["normal_trading_m1_p1"] == 1)
                        & (panel["no_limit_m1_p1"] == 1)
                        & (panel["obs_car_m1_p1"] == 3)
                    ).sum()
                )
                if not panel.empty
                else 0,
            },
        ]
    )
    summary.to_csv(OUT_DIR / "v4_csmar_return_coverage_summary.csv", index=False)
    if suffix:
        summary.to_csv(OUT_DIR / f"v4_csmar_return_coverage_summary{suffix}.csv", index=False)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--refresh-cache", action="store_true")
    parser.add_argument("--max-peer-rank", type=int, default=MAX_PEER_RANK)
    parser.add_argument("--index-code", default="000001")
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    suffix = f"_top{args.max_peer_rank}"
    x = read_x_sample(args.max_peer_rank)
    target_codes = set(x["peer_code"].dropna().unique())
    print(
        f"Loaded X sample: {len(x):,} peer-event obs, "
        f"{x['event_id'].nunique():,} events, {len(target_codes):,} peer firms",
        flush=True,
    )

    stock = build_stock_return_cache(target_codes, refresh=args.refresh_cache)
    mkt = build_market_return_cache(refresh=args.refresh_cache, index_code=args.index_code)
    stock.to_csv(OUT_DIR / "v4_csmar_peer_daily_returns_subset.csv", index=False)
    mkt.to_csv(OUT_DIR / f"v4_csmar_market_returns_{args.index_code}.csv", index=False)

    panel = compute_event_returns(x, stock, mkt)
    if panel.empty:
        extra_cols = [
            "event_date_for_car",
            "estimation_obs",
            "peer_alpha",
            "peer_beta",
            "peer_ar0",
            "peer_car_0_p1",
            "peer_car_m1_p1",
            "peer_abs_ar0",
            "peer_abs_car_0_p1",
            "peer_abs_car_m1_p1",
            "peer_raw_ret_0",
            "peer_raw_ret_0_p1",
            "peer_raw_ret_m1_p1",
            "mkt_ret_0",
            "mkt_ret_0_p1",
            "mkt_ret_m1_p1",
            "obs_ar0",
            "obs_car_0_p1",
            "obs_car_m1_p1",
            "normal_trading_ar0",
            "normal_trading_0_p1",
            "normal_trading_m1_p1",
            "no_limit_ar0",
            "no_limit_0_p1",
            "no_limit_m1_p1",
        ]
        panel = pd.DataFrame(columns=list(x.columns) + extra_cols)
    panel_path = OUT_DIR / "v4_peer_event_with_car_csmar.csv"
    panel.to_csv(panel_path, index=False)
    panel.to_csv(OUT_DIR / f"v4_peer_event_with_car_csmar{suffix}.csv", index=False)
    write_summary(x, stock, mkt, panel, suffix=suffix)

    reg = run_regressions(panel)
    reg.to_csv(OUT_DIR / "v4_csmar_regression_summary.csv", index=False)
    reg.to_csv(OUT_DIR / f"v4_csmar_regression_summary{suffix}.csv", index=False)

    print(f"Wrote panel: {panel_path}", flush=True)
    print(reg.to_string(index=False), flush=True)


if __name__ == "__main__":
    main()
