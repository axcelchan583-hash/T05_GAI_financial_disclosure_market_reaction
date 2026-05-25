#!/usr/bin/env python3
"""Pilot for v3: GenAI disclosure specificity -> |CAR[-1,+1]|.

This is a feasibility run, not a final event study. It:
1. Queries CNINFO for GenAI/AI-related announcements.
2. Keeps first non-periodic A-share announcement events with GenAI text.
3. Downloads PDFs and extracts text.
4. Builds a transparent rule-based specificity measure for GenAI spans.
5. Links CSMAR daily returns and computes market-model CAR.
"""

from __future__ import annotations

import csv
import html
import json
import math
import re
import subprocess
import time
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlencode, urljoin
from urllib.request import Request, urlopen

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "results" / "v3_genai_specificity_car_pilot"
DOC_OUT = ROOT / "docs" / "选题" / "18_v3_genai_specificity_car_pilot_20260522.md"
RAW_DIR = ROOT / "data" / "raw" / "cninfo_genai_announcements_20260522"
RAW_JSON_DIR = RAW_DIR / "raw_json"
PDF_DIR = RAW_DIR / "raw_pdf"
TXT_DIR = RAW_DIR / "text"

DATA_ROOT = Path("/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/上市公司财务信息")
DAILY_DIR = DATA_ROOT / "_daily_2021_2026"
INDEX_XLSX = DATA_ROOT / "TRD_Index.xlsx"

QUERY_URL = "http://www.cninfo.com.cn/new/hisAnnouncement/query"
STATIC_BASE = "https://static.cninfo.com.cn/"
SEARCH_PAGE = "http://www.cninfo.com.cn/new/commonUrl/pageOfSearch?url=disclosure/list/search"

START_DATE = "2023-01-01"
END_DATE = "2026-05-21"
PAGE_SIZE = 30

QUERY_TERMS = [
    "大模型",
    "大语言模型",
    "生成式人工智能",
    "生成式AI",
    "AIGC",
    "ChatGPT",
    "DeepSeek",
    "智能体",
    "模型备案",
    "人工智能大模型",
    "人工智能",
]

GENAI_TERMS = [
    "生成式人工智能",
    "生成式AI",
    "生成式 AI",
    "AIGC",
    "大模型",
    "大语言模型",
    "语言大模型",
    "基础模型",
    "预训练模型",
    "多模态大模型",
    "ChatGPT",
    "GPT",
    "DeepSeek",
    "文心一言",
    "通义千问",
    "通义",
    "盘古大模型",
    "混元",
    "豆包",
    "智能体",
    "模型备案",
]

EXCLUDE_TITLE_PAT = re.compile(
    r"年度报告|年报|半年度报告|半年报|季度报告|一季度报告|三季度报告|"
    r"业绩预告|业绩快报|业绩说明会|投资者关系|调研活动|机构调研|"
    r"问询函|回复函|审核问询|核查意见|法律意见书|保荐|独立财务顾问|"
    r"摘要|审计报告|财务报表",
    re.I,
)

A_SHARE_PAT = re.compile(r"^(000|001|002|003|300|301|600|601|603|605|688)\d{3}$")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "Referer": SEARCH_PAGE,
    "Origin": "http://www.cninfo.com.cn",
}


def strip_tags(text: str) -> str:
    text = html.unescape(text or "")
    text = re.sub(r"<[^>]+>", "", text)
    return re.sub(r"\s+", "", text).strip()


def safe_name(text: str) -> str:
    text = re.sub(r"[\\/:*?\"<>|]", "_", text)
    text = re.sub(r"\s+", "_", text)
    return text[:180].strip("_")


def ts_to_date(ms: object) -> str:
    try:
        return datetime.fromtimestamp(int(ms) / 1000).strftime("%Y-%m-%d")
    except Exception:
        return ""


def http_post_json(url: str, data: dict[str, object], timeout: int = 25) -> dict:
    payload = urlencode(data).encode("utf-8")
    headers = {
        **HEADERS,
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    }
    req = Request(url, data=payload, headers=headers, method="POST")
    with urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8", errors="replace"))


def http_get_bytes(url: str, timeout: int = 35) -> tuple[int, bytes]:
    req = Request(url, headers={"User-Agent": HEADERS["User-Agent"]}, method="GET")
    with urlopen(req, timeout=timeout) as resp:
        return int(getattr(resp, "status", 200)), resp.read()


def query_cninfo(searchkey: str, page_num: int) -> dict:
    data = {
        "pageNum": page_num,
        "pageSize": PAGE_SIZE,
        "column": "szse",
        "tabName": "fulltext",
        "plate": "",
        "stock": "",
        "searchkey": searchkey,
        "secid": "",
        "category": "",
        "trade": "",
        "seDate": f"{START_DATE}~{END_DATE}",
        "sortName": "",
        "sortType": "",
        "isHLtitle": "true",
    }
    for attempt in range(4):
        try:
            return http_post_json(QUERY_URL, data, timeout=25)
        except Exception:
            if attempt == 3:
                raise
            time.sleep(1 + attempt)
    raise RuntimeError("unreachable")


def query_all_terms() -> list[dict[str, str]]:
    RAW_JSON_DIR.mkdir(parents=True, exist_ok=True)
    rows: dict[str, dict[str, str]] = {}
    query_counts = []
    for term in QUERY_TERMS:
        first = query_cninfo(term, 1)
        total = int(first.get("totalRecordNum") or 0)
        pages = max(1, math.ceil(total / PAGE_SIZE))
        query_counts.append({"query_term": term, "total": total, "pages": pages})
        print(f"CNINFO {term}: total={total}, pages={pages}", flush=True)
        for page, payload in [(1, first)]:
            (RAW_JSON_DIR / f"{safe_name(term)}_page_{page:04d}.json").write_text(
                json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
            )
        for page in range(2, pages + 1):
            time.sleep(0.15)
            payload = query_cninfo(term, page)
            (RAW_JSON_DIR / f"{safe_name(term)}_page_{page:04d}.json").write_text(
                json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            for row in payload.get("announcements") or []:
                nr = normalize_announcement(row, term)
                if nr["announcement_id"]:
                    rows[nr["announcement_id"]] = nr
        for row in first.get("announcements") or []:
            nr = normalize_announcement(row, term)
            if nr["announcement_id"]:
                rows.setdefault(nr["announcement_id"], nr)
                if term not in rows[nr["announcement_id"]]["query_terms"].split(";"):
                    terms = set(rows[nr["announcement_id"]]["query_terms"].split(";")) | {term}
                    rows[nr["announcement_id"]]["query_terms"] = ";".join(sorted(t for t in terms if t))
    write_csv(OUT_DIR / "cninfo_query_counts.csv", query_counts, ["query_term", "total", "pages"])
    return sorted(rows.values(), key=lambda r: (r["announcement_date"], r["sec_code"], r["announcement_id"]))


def normalize_announcement(row: dict, term: str) -> dict[str, str]:
    adjunct = row.get("adjunctUrl") or ""
    return {
        "query_terms": term,
        "sec_code": str(row.get("secCode") or "").zfill(6),
        "sec_name": strip_tags(row.get("secName") or ""),
        "org_id": row.get("orgId") or "",
        "announcement_id": str(row.get("announcementId") or ""),
        "announcement_title": strip_tags(row.get("announcementTitle") or ""),
        "announcement_date": ts_to_date(row.get("announcementTime")),
        "adjunct_url": adjunct,
        "pdf_url": urljoin(STATIC_BASE, adjunct) if adjunct else "",
        "adjunct_size": str(row.get("adjunctSize") or ""),
        "adjunct_type": row.get("adjunctType") or "",
        "page_column": row.get("pageColumn") or "",
    }


def write_csv(path: Path, rows: list[dict], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def is_candidate(row: dict[str, str]) -> bool:
    code = row["sec_code"]
    title = row["announcement_title"]
    if not A_SHARE_PAT.match(code):
        return False
    if EXCLUDE_TITLE_PAT.search(title):
        return False
    if not row["pdf_url"]:
        return False
    return True


def download_and_extract(row: dict[str, str]) -> dict[str, str]:
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    TXT_DIR.mkdir(parents=True, exist_ok=True)
    filename = safe_name(f"{row['announcement_date']}_{row['sec_code']}_{row['sec_name']}_{row['announcement_id']}_{row['announcement_title']}")
    pdf_path = PDF_DIR / f"{filename}.pdf"
    txt_path = TXT_DIR / f"{filename}.txt"
    out = {**row, "pdf_file": "", "txt_file": "", "download_status": ""}
    try:
        if not pdf_path.exists():
            status, content = http_get_bytes(row["pdf_url"], timeout=35)
            if status != 200 or b"%PDF" not in content[:1024]:
                out["download_status"] = f"download_failed_{status}"
                return out
            pdf_path.write_bytes(content)
            time.sleep(0.05)
        if not txt_path.exists():
            proc = subprocess.run(["pdftotext", "-layout", str(pdf_path), str(txt_path)], capture_output=True, text=True)
            if proc.returncode != 0:
                out["download_status"] = "pdftotext_failed"
                out["pdf_file"] = str(pdf_path.relative_to(ROOT))
                return out
        out["pdf_file"] = str(pdf_path.relative_to(ROOT))
        out["txt_file"] = str(txt_path.relative_to(ROOT))
        out["download_status"] = "ok"
        return out
    except Exception as exc:
        out["download_status"] = f"error_{type(exc).__name__}"
        return out


def split_sentences(text: str) -> list[str]:
    text = re.sub(r"\s+", " ", text)
    parts = re.split(r"(?<=[。！？；;])", text)
    return [p.strip() for p in parts if len(p.strip()) >= 8]


def extract_genai_span(text: str) -> tuple[str, str, str]:
    sentences = split_sentences(text)
    matched = []
    terms_found = set()
    for sent in sentences:
        terms = [t for t in GENAI_TERMS if t.lower() in sent.lower()]
        if terms:
            matched.append(sent[:1000])
            terms_found.update(terms)
    return " ".join(matched), ";".join(sorted(terms_found)), " || ".join(matched[:5])


def specificity_metrics(span: str) -> dict[str, float | int | str]:
    clean = re.sub(r"\s+", "", span)
    token_count = max(1, len(re.findall(r"[\u4e00-\u9fff]|[A-Za-z0-9]+", clean)))
    patterns = {
        "dates": r"(?:20[2-9]\d年\d{0,2}月?\d{0,2}日?|20[2-9]\d[-/]\d{1,2}[-/]\d{1,2}|[一二三四1234]季度|Q[1-4])",
        "money": r"\d+(?:\.\d+)?(?:亿元|万元|元|万美元|美元|人民币)",
        "percent": r"\d+(?:\.\d+)?%",
        "numbers": r"\d+(?:\.\d+)?(?:个|项|款|台|套|人|家|年|月|日|亿|万)?",
        "quoted_names": r"[“《][^”》]{2,30}[”》]",
        "orgs": r"[\u4e00-\u9fa5A-Za-z0-9]{2,30}(?:公司|集团|大学|研究院|实验室|中心|平台|科技|智能|网络)",
        "partners": r"(?:华为|百度|阿里|腾讯|科大讯飞|智谱|商汤|字节|DeepSeek|OpenAI|微软|英伟达|通义|文心)",
        "stage": r"(?:已上线|已发布|已备案|已通过|已落地|已部署|签署|合作|共建|推出|发布|升级|备案通过)",
    }
    counts = {k: len(re.findall(pat, clean, flags=re.I)) for k, pat in patterns.items()}
    specific_items = sum(counts.values())
    return {
        **counts,
        "specific_items": specific_items,
        "genai_span_tokens": token_count,
        "specificity": 100 * specific_items / token_count,
        "genai_mentions": sum(clean.lower().count(t.lower()) for t in GENAI_TERMS),
    }


def build_event_candidates(rows: list[dict[str, str]]) -> pd.DataFrame:
    fields = list(rows[0].keys()) if rows else []
    write_csv(OUT_DIR / "cninfo_genai_raw_hits.csv", rows, fields)
    filtered = [r for r in rows if is_candidate(r)]
    write_csv(OUT_DIR / "cninfo_genai_title_filtered_candidates.csv", filtered, fields)
    downloaded = [download_and_extract(r) for r in filtered]
    write_csv(OUT_DIR / "cninfo_genai_downloaded_candidates.csv", downloaded, list(downloaded[0].keys()) if downloaded else fields)

    event_rows = []
    for row in downloaded:
        if row.get("download_status") != "ok" or not row.get("txt_file"):
            continue
        txt = (ROOT / row["txt_file"]).read_text(encoding="utf-8", errors="ignore")
        span, matched_terms, examples = extract_genai_span(txt)
        if not matched_terms:
            continue
        metrics = specificity_metrics(span)
        event_rows.append({**row, **metrics, "matched_genai_terms": matched_terms, "genai_span_examples": examples})
    df = pd.DataFrame(event_rows)
    if df.empty:
        return df
    df["announcement_date_dt"] = pd.to_datetime(df["announcement_date"], errors="coerce")
    df = df.sort_values(["sec_code", "announcement_date_dt", "announcement_id"])
    df["event_rank_within_firm"] = df.groupby("sec_code").cumcount() + 1
    df_first = df[df["event_rank_within_firm"] == 1].copy()
    df.to_csv(OUT_DIR / "cninfo_genai_event_candidates_with_specificity.csv", index=False, encoding="utf-8-sig")
    df_first.to_csv(OUT_DIR / "cninfo_genai_first_events_with_specificity.csv", index=False, encoding="utf-8-sig")
    return df_first


def excel_rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    ws.reset_dimensions()
    for row in ws.iter_rows(values_only=True):
        yield row
    wb.close()


def load_market_returns() -> pd.DataFrame:
    rows = []
    for i, row in enumerate(excel_rows(INDEX_XLSX)):
        if i < 3:
            continue
        if str(row[0]).zfill(6) != "000001":
            continue
        if row[1] is None or row[7] is None:
            continue
        rows.append({"date": row[1], "mkt_ret": float(row[7])})
    mkt = pd.DataFrame(rows)
    mkt["date"] = pd.to_datetime(mkt["date"])
    return mkt.sort_values("date")


def load_stock_returns(symbols: set[str], start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    rows = []
    files = sorted(DAILY_DIR.glob("TRD_Dalyr*.xlsx"))
    print(f"Loading daily returns for {len(symbols)} symbols from {len(files)} files", flush=True)
    for path in files:
        t0 = time.time()
        hit = 0
        for i, row in enumerate(excel_rows(path)):
            if i < 3:
                continue
            code = str(row[0]).zfill(6)
            if code not in symbols:
                continue
            d = pd.to_datetime(row[1], errors="coerce")
            if pd.isna(d) or d < start or d > end:
                continue
            ret = row[10] if row[10] is not None else row[11]
            if ret is None:
                continue
            rows.append({"sec_code": code, "date": d, "ret": float(ret), "trdsta": row[16], "markettype": row[14]})
            hit += 1
        print(f"  {path.name}: hits={hit}, elapsed={time.time()-t0:.1f}s", flush=True)
    return pd.DataFrame(rows).sort_values(["sec_code", "date"])


def compute_car(events: pd.DataFrame) -> pd.DataFrame:
    if events.empty:
        return events
    mkt = load_market_returns()
    last_mkt = mkt["date"].max()
    events = events.copy()
    events["raw_event_date"] = pd.to_datetime(events["announcement_date"], errors="coerce")
    events = events[events["raw_event_date"].notna() & (events["raw_event_date"] <= last_mkt)].copy()
    if events.empty:
        return events
    start = events["raw_event_date"].min() - timedelta(days=380)
    end = events["raw_event_date"].max() + timedelta(days=10)
    stock = load_stock_returns(set(events["sec_code"]), start, end)
    if stock.empty:
        return events
    ret = stock.merge(mkt, on="date", how="left").dropna(subset=["mkt_ret", "ret"])
    trading_dates = sorted(mkt["date"].unique())
    date_index = {d: i for i, d in enumerate(trading_dates)}

    out = []
    for _, ev in events.iterrows():
        code = ev["sec_code"]
        raw_date = ev["raw_event_date"]
        possible = [d for d in trading_dates if d >= raw_date]
        if not possible:
            continue
        event_date = possible[0]
        idx = date_index[event_date]
        est_dates = set(trading_dates[max(0, idx - 220): max(0, idx - 20)])
        win_dates = set(trading_dates[max(0, idx - 1): min(len(trading_dates), idx + 2)])
        firm_ret = ret[ret["sec_code"] == code].copy()
        est = firm_ret[firm_ret["date"].isin(est_dates)]
        win = firm_ret[firm_ret["date"].isin(win_dates)]
        if len(est) < 80 or len(win) < 2:
            continue
        x = np.column_stack([np.ones(len(est)), est["mkt_ret"].to_numpy()])
        y = est["ret"].to_numpy()
        alpha, beta = np.linalg.lstsq(x, y, rcond=None)[0]
        win = win.copy()
        win["abret"] = win["ret"] - (alpha + beta * win["mkt_ret"])
        car = float(win["abret"].sum())
        out.append(
            {
                **ev.to_dict(),
                "event_date_for_car": pd.Timestamp(event_date).strftime("%Y-%m-%d"),
                "estimation_obs": len(est),
                "event_window_obs": len(win),
                "alpha": alpha,
                "beta": beta,
                "car_m1_p1": car,
                "abs_car_m1_p1": abs(car),
                "raw_return_m1_p1": float(win["ret"].sum()),
                "mkt_return_m1_p1": float(win["mkt_ret"].sum()),
            }
        )
    return pd.DataFrame(out)


def run_regression(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or len(df) < 8:
        return pd.DataFrame([{"note": "too few CAR-linked events", "n": len(df)}])
    data = df[["abs_car_m1_p1", "specificity", "genai_span_tokens", "genai_mentions"]].dropna().copy()
    data["log_span_tokens"] = np.log1p(data["genai_span_tokens"].astype(float))
    y = data["abs_car_m1_p1"].astype(float).to_numpy()
    xcols = ["specificity", "log_span_tokens", "genai_mentions"]
    x = np.column_stack([np.ones(len(data))] + [data[c].astype(float).to_numpy() for c in xcols])
    terms = ["Intercept"] + xcols
    beta = np.linalg.lstsq(x, y, rcond=None)[0]
    resid = y - x @ beta
    n, k = x.shape
    sigma2 = float((resid @ resid) / max(n - k, 1))
    vcov = sigma2 * np.linalg.pinv(x.T @ x)
    se = np.sqrt(np.diag(vcov))
    rows = []
    for term, coef, stderr in zip(terms, beta, se):
        rows.append({"term": term, "coef": coef, "std_err": stderr, "nobs": n, "note": "pilot OLS, no FE, classical SE"})
    return pd.DataFrame(rows)


def write_report(raw: pd.DataFrame, first: pd.DataFrame, car: pd.DataFrame, reg: pd.DataFrame) -> None:
    def table(df: pd.DataFrame, max_rows: int = 20) -> str:
        if df.empty:
            return "无"
        show = df.head(max_rows)
        cols = list(show.columns)
        lines = ["| " + " | ".join(cols) + " |", "| " + " | ".join(["---"] * len(cols)) + " |"]
        for _, row in show.iterrows():
            vals = [str(row.get(c, "")).replace("\n", " ").replace("|", "\\|")[:220] for c in cols]
            lines.append("| " + " | ".join(vals) + " |")
        return "\n".join(lines)

    summary = {
        "raw_hits": len(raw),
        "first_events_with_specificity": len(first),
        "car_linked_events": len(car),
        "car_mean": car["abs_car_m1_p1"].mean() if not car.empty else np.nan,
        "specificity_mean": first["specificity"].mean() if not first.empty else np.nan,
    }
    examples_cols = [
        "sec_code",
        "sec_name",
        "announcement_date",
        "announcement_title",
        "specificity",
        "genai_span_tokens",
        "matched_genai_terms",
        "abs_car_m1_p1",
        "car_m1_p1",
    ]
    report = f"""# v3 最小试跑：GenAI 披露具体性 -> |CAR|

日期：2026-05-22

## 1. 这次跑了什么

这是 v3 主线的最小可行试跑：

```text
Main X = GenAI disclosure specificity / Level of Detail
Main Y = |CAR[-1,+1]|
```

本轮使用巨潮公告关键词检索构造非定期 GenAI 公告样本，并用 CSMAR 日个股回报率与上证指数回报率估计市场模型 CAR。

注意：这是粗 pilot。公告时点暂按公告日期处理，尚未区分盘后披露；specificity 是规则化近似指标，还不是最终 Hope-style 中文 NER 版本。

## 2. 样本概况

| item | value |
|---|---:|
| raw CNINFO hits | {summary['raw_hits']} |
| first events with GenAI text and specificity | {summary['first_events_with_specificity']} |
| CAR-linked first events | {summary['car_linked_events']} |
| mean specificity | {summary['specificity_mean']:.4f} |
| mean |CAR[-1,+1]| | {summary['car_mean']:.4f} |

## 3. 粗回归

{table(reg)}

## 4. CAR-linked examples

{table(car[[c for c in examples_cols if c in car.columns]].sort_values('abs_car_m1_p1', ascending=False) if not car.empty else car, max_rows=20)}

## 5. 初步判断

这轮只回答“能不能跑通”。若 CAR-linked 样本量太小，下一步需要扩大事件库，而不是解释系数。

关键改进：

1. 用全部巨潮公告库或更稳的公告下载源扩展事件样本。
2. 明确盘后披露顺延规则。
3. 用正式中文 NER 固化 Hope-style specificity。
4. 加入同日重大公告污染过滤。
5. 再接 ABVOL、signed CAR 和 PriorAICapability 交互。
"""
    DOC_OUT.parent.mkdir(parents=True, exist_ok=True)
    DOC_OUT.write_text(report, encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for p in [RAW_JSON_DIR, PDF_DIR, TXT_DIR]:
        p.mkdir(parents=True, exist_ok=True)
    raw_rows = query_all_terms()
    raw = pd.DataFrame(raw_rows)
    first = build_event_candidates(raw_rows)
    car = compute_car(first)
    if not car.empty:
        car.to_csv(OUT_DIR / "first_events_with_car.csv", index=False, encoding="utf-8-sig")
    reg = run_regression(car)
    reg.to_csv(OUT_DIR / "pilot_ols_abs_car_on_specificity.csv", index=False, encoding="utf-8-sig")
    write_report(raw, first, car, reg)
    print(f"raw_hits={len(raw)} first_events={len(first)} car_linked={len(car)}", flush=True)
    print(f"report={DOC_OUT}", flush=True)


if __name__ == "__main__":
    main()
