#!/usr/bin/env python3
"""Extract the local CSMAR algorithm-info workbook into interim CSV.

The source workbook is proprietary/local and is not copied into this repo.
Set CSMAR_ALGORITHM_ZIP to override the default local path.
"""

from __future__ import annotations

import csv
import os
import re
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ZIP = Path(
    "/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/"
    "算法治理/互联网信息服务算法信息表145714958(仅供沪江大学使用).zip"
)
OUT_DIR = ROOT / "data" / "interim"
OUT_CSV = OUT_DIR / "csmar_algorithm_info_records.csv"
SUMMARY_CSV = OUT_DIR / "csmar_algorithm_info_summary_by_month.csv"

FIELD_MAP = {
    "SgnMonth": "source_batch",
    "InstitutionName": "filing_entity",
    "InstitutionID": "institution_id",
    "AlgorithmName": "algorithm_name",
    "AlgorithmType": "algorithm_type",
    "ApplicationProduct": "application_product",
    "MainPurpose": "main_purpose",
    "BasicPrinciples": "basic_principles",
    "OperatingMechanism": "operating_mechanism",
    "ApplicationScenarios": "application_scenarios",
    "PurposeIntention": "purpose_intention",
    "RecordNumber": "record_no",
    "DistributionDate": "distribution_date",
    "Comments": "comments",
}


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def source_zip_path() -> Path:
    return Path(os.environ.get("CSMAR_ALGORITHM_ZIP", str(DEFAULT_ZIP))).expanduser()


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    zip_path = source_zip_path()
    if not zip_path.exists():
        raise FileNotFoundError(f"CSMAR algorithm zip not found: {zip_path}")

    with zipfile.ZipFile(zip_path) as zf:
        with zf.open("AI_AlgorithmInfo.xlsx") as f:
            workbook_bytes = f.read()

    # This CSMAR workbook exposes only the first column under openpyxl's
    # read_only mode, so load the normal worksheet object for reliable columns.
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=False, data_only=True)
    sheet = workbook["sheet1"]
    raw_headers = [clean_cell(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    headers = [FIELD_MAP.get(header, header) for header in raw_headers]

    rows: list[dict[str, str]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        row = {headers[idx]: clean_cell(value) for idx, value in enumerate(raw_row)}
        record_no = row.get("record_no", "")
        if "网信算备" not in record_no:
            continue
        row["source_workbook"] = "AI_AlgorithmInfo.xlsx"
        row["source_zip_basename"] = zip_path.name
        rows.append(row)

    fields = list(FIELD_MAP.values()) + ["source_workbook", "source_zip_basename"]
    write_csv(OUT_CSV, rows, fields)

    summary: dict[tuple[str, str], int] = {}
    for row in rows:
        key = (row["source_batch"], row["algorithm_type"])
        summary[key] = summary.get(key, 0) + 1
    summary_rows = [
        {"source_batch": batch, "algorithm_type": algorithm_type, "n": n}
        for (batch, algorithm_type), n in sorted(summary.items())
    ]
    write_csv(SUMMARY_CSV, summary_rows, ["source_batch", "algorithm_type", "n"])

    print(f"Extracted {len(rows)} records into {OUT_CSV}")


if __name__ == "__main__":
    main()
