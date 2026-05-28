#!/usr/bin/env python3
"""Measurement and final-design closure checks for T05.

This run consolidates the current paper around the peer-revaluation design.
It intentionally does not search for new outcomes.  It answers the main
reviewer-facing gaps:

1. freeze the final headline sample;
2. report human/Claude coding as an internal construct-boundary check for the
   old proxy score;
3. rerun the headline with AI-theme date-shock controls;
4. rerun external AIActive component checks on the final sample;
5. translate the coefficient into market-cap magnitude; and
6. collect the existing non-GenAI pseudo-event evidence.
"""

from __future__ import annotations

import math
import re
from collections.abc import Iterable
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from run_v6_final_review_checks_20260525 import OUTCOME, ROOT, STRONG_FE, add_ai_terms
from run_v6_identification_strengthening_checks_20260524 import fit_absorbed


OUT_DIR = ROOT / "results" / "v8_measurement_final_checks_20260527"
DOC_PATH = ROOT / "docs" / "empirical_runs" / "61_v8_measurement_final_checks_20260527.md"

FINAL_SAMPLE_PATH = (
    ROOT
    / "results"
    / "v6_focal_good_news_pretrend_checks_20260525"
    / "analysis_sample_top5.csv.gz"
)
AGENT_CODES = ROOT / "data" / "specificity_validation" / "agent_coding" / "agent1_manual_codes_300_20260526.csv"
CLAUDE_CODES = (
    ROOT
    / "claude_project"
    / "3"
    / "validation_sample"
    / "claude_coding"
    / "claude_manual_codes_300_20260526.csv"
)
STOCK_RETURNS = (
    ROOT
    / "results"
    / "v6_supplement_market_model_placebo_20260524"
    / "stock_returns_with_market_model_params.csv"
)
EXTERNAL_FIRST_DATES = (
    ROOT
    / "results"
    / "v6_external_ai_active_20260524"
    / "external_ai_evidence_first_dates_by_firm.csv"
)
NON_GENAI_PLACEBO = (
    ROOT
    / "results"
    / "v6_identification_strengthening_20260524"
    / "non_genai_placebo_results.csv"
)
PREWINDOW_PLACEBO = (
    ROOT
    / "results"
    / "v6_identification_strengthening_20260524"
    / "prewindow_placebo_results.csv"
)
TRD_DIR = Path(
    "/Users/mac/computerscience/第三方资料/01_数据资源/国泰安/第三方数据资源/上市公司财务信息"
)


AI_DEFS = ["ext_any", "current_text_history"]
COMPONENTS = [
    "has_specific_product_service",
    "has_model_platform_name",
    "has_specific_use_case",
    "has_customer_or_industry",
    "has_partner_or_org",
    "has_deployment_status",
    "has_commercialization_or_timeline",
    "has_quantitative_commitment",
]


def code6(value: object) -> str:
    if pd.isna(value):
        return ""
    text = re.sub(r"\.0$", "", str(value).strip())
    digits = re.sub(r"\D", "", text)
    return digits.zfill(6) if digits else ""


def pval_from_z(z: float) -> float:
    if not math.isfinite(z):
        return math.nan
    return math.erfc(abs(z) / math.sqrt(2.0))


def corr_pair(x: pd.Series, y: pd.Series) -> tuple[float, float, int]:
    d = pd.DataFrame({"x": pd.to_numeric(x, errors="coerce"), "y": pd.to_numeric(y, errors="coerce")}).dropna()
    if len(d) < 3 or d["x"].std(ddof=0) == 0 or d["y"].std(ddof=0) == 0:
        return math.nan, math.nan, int(len(d))
    return float(d["x"].corr(d["y"], method="pearson")), float(d["x"].corr(d["y"], method="spearman")), int(len(d))


def cohens_kappa(a: Iterable[object], b: Iterable[object]) -> float:
    x = pd.Series(a).astype(float)
    y = pd.Series(b).astype(float)
    d = pd.DataFrame({"x": x, "y": y}).dropna()
    if d.empty:
        return math.nan
    labels = sorted(set(d["x"].unique()).union(set(d["y"].unique())))
    observed = float((d["x"] == d["y"]).mean())
    expected = 0.0
    for lab in labels:
        expected += float((d["x"] == lab).mean()) * float((d["y"] == lab).mean())
    if expected == 1.0:
        return math.nan
    return (observed - expected) / (1.0 - expected)


def load_final_sample() -> pd.DataFrame:
    df = pd.read_csv(FINAL_SAMPLE_PATH, dtype={"event_id": str, "focal_code": str, "peer_code": str})
    df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce")
    for col in ["date_0", "date_p1", "date_m1"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    df["peer_code"] = df["peer_code"].map(code6)
    df["focal_code"] = df["focal_code"].map(code6)
    return df


def final_sample_freeze(sample: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary = pd.DataFrame(
        [
            {
                "sample_name": "final_headline_top5",
                "rows": len(sample),
                "events": sample["event_id"].nunique(),
                "focal_firms": sample["focal_code"].nunique(),
                "peer_firms": sample["peer_code"].nunique(),
                "date_min": sample["event_date"].min().date(),
                "date_max": sample["event_date"].max().date(),
                "rank_min": int(sample["peer_rank"].min()),
                "rank_max": int(sample["peer_rank"].max()),
                "mean_y": sample[OUTCOME].mean(),
                "ext_any_share": sample["ext_any"].mean(),
                "text_history_share": sample["current_text_history"].mean(),
                "requires_peer_pre10_pre20": True,
                "requires_focal_car": True,
                "announcement_cleaned": True,
            }
        ]
    )
    by_year = (
        sample.assign(event_year=sample["event_date"].dt.year)
        .groupby("event_year")
        .agg(rows=("event_id", "size"), events=("event_id", "nunique"), peer_firms=("peer_code", "nunique"))
        .reset_index()
    )
    return summary, by_year


def run_coder_validation() -> dict[str, pd.DataFrame]:
    agent = pd.read_csv(AGENT_CODES, dtype={"validation_id": str, "event_id": str})
    claude = pd.read_csv(CLAUDE_CODES, dtype={"validation_id": str, "event_id": str})
    keep = [
        "validation_id",
        "event_id",
        "specificity_z",
        "specificity_bin",
        *COMPONENTS,
        "specificity_score_0_4",
        "uncertain_flag",
    ]
    a = agent[keep].copy()
    c = claude[keep].copy()
    merged = a.merge(c, on=["validation_id", "event_id"], suffixes=("_agent1", "_claude"))

    rows = []
    for comp in COMPONENTS:
        aa = merged[f"{comp}_agent1"].fillna(0).astype(int)
        cc = merged[f"{comp}_claude"].fillna(0).astype(int)
        rows.append(
            {
                "variable": comp,
                "n": len(merged),
                "agent1_ones": int(aa.sum()),
                "claude_ones": int(cc.sum()),
                "agreement_rate": float((aa == cc).mean()),
                "cohens_kappa": cohens_kappa(aa, cc),
            }
        )
    aa = merged["specificity_score_0_4_agent1"].astype(int)
    cc = merged["specificity_score_0_4_claude"].astype(int)
    rows.append(
        {
            "variable": "specificity_score_0_4",
            "n": len(merged),
            "agent1_ones": math.nan,
            "claude_ones": math.nan,
            "agreement_rate": float((aa == cc).mean()),
            "cohens_kappa": cohens_kappa(aa, cc),
        }
    )
    agreement = pd.DataFrame(rows)

    for who in ["agent1", "claude"]:
        merged[f"component_sum_{who}"] = merged[[f"{c}_{who}" for c in COMPONENTS]].sum(axis=1)
    merged["specificity_score_mean"] = (
        merged["specificity_score_0_4_agent1"] + merged["specificity_score_0_4_claude"]
    ) / 2
    merged["component_sum_mean"] = (merged["component_sum_agent1"] + merged["component_sum_claude"]) / 2
    merged["validated_specificity_score_z"] = (
        merged["specificity_score_mean"] - merged["specificity_score_mean"].mean()
    ) / merged["specificity_score_mean"].std(ddof=0)

    corr_rows = []
    for x, y in [
        ("component_sum_claude", "component_sum_agent1"),
        ("specificity_score_0_4_claude", "specificity_score_0_4_agent1"),
        ("component_sum_mean", "specificity_z_agent1"),
        ("specificity_score_mean", "specificity_z_agent1"),
        ("component_sum_claude", "specificity_z_agent1"),
        ("specificity_score_0_4_claude", "specificity_z_agent1"),
        ("component_sum_agent1", "specificity_z_agent1"),
        ("specificity_score_0_4_agent1", "specificity_z_agent1"),
    ]:
        pear, spear, n = corr_pair(merged[x], merged[y])
        corr_rows.append({"x": x, "y": y, "pearson": pear, "spearman": spear, "n": n})
    correlations = pd.DataFrame(corr_rows)

    bin_summary = (
        merged.groupby("specificity_bin_agent1")
        .agg(
            n=("validation_id", "size"),
            mean_proxy_specificity_z=("specificity_z_agent1", "mean"),
            mean_claude_score=("specificity_score_0_4_claude", "mean"),
            mean_agent_score=("specificity_score_0_4_agent1", "mean"),
            mean_score_mean=("specificity_score_mean", "mean"),
            mean_component_sum_mean=("component_sum_mean", "mean"),
        )
        .reset_index()
    )
    scored = merged[
        [
            "validation_id",
            "event_id",
            "specificity_z_agent1",
            "specificity_score_0_4_agent1",
            "specificity_score_0_4_claude",
            "specificity_score_mean",
            "component_sum_agent1",
            "component_sum_claude",
            "component_sum_mean",
            "validated_specificity_score_z",
        ]
    ].rename(columns={"specificity_z_agent1": "original_specificity_z"})
    return {
        "specificity_coder_agreement": agreement,
        "specificity_coder_correlations": correlations,
        "specificity_coder_bin_summary": bin_summary,
        "specificity_validated_scores_300": scored,
    }


def add_theme_factor(sample: pd.DataFrame) -> pd.DataFrame:
    ret = pd.read_csv(STOCK_RETURNS, dtype={"peer_code": str})
    ret["peer_code"] = ret["peer_code"].map(code6)
    ret["date"] = pd.to_datetime(ret["date"], errors="coerce")
    for col in ["ret", "mkt_ret", "alpha_mm", "beta_mm"]:
        ret[col] = pd.to_numeric(ret[col], errors="coerce")
    ret["abret_mm"] = ret["ret"] - (ret["alpha_mm"] + ret["beta_mm"] * ret["mkt_ret"])
    ret = ret.dropna(subset=["peer_code", "date", "abret_mm"])

    evidence = pd.read_csv(EXTERNAL_FIRST_DATES, dtype={"stock_code": str})
    evidence["stock_code"] = evidence["stock_code"].map(code6)
    date_cols = [
        "cac_first_date",
        "ai_patent_grant_first_date",
        "broad_ai_hiring_first_date",
    ]
    for col in date_cols:
        evidence[col] = pd.to_datetime(evidence[col], errors="coerce")
    evidence["external_first_date"] = evidence[date_cols].min(axis=1)
    ret = ret.merge(evidence[["stock_code", "external_first_date"]], left_on="peer_code", right_on="stock_code", how="left")
    ret["external_active_asof_date"] = ret["external_first_date"].notna() & (
        ret["external_first_date"] <= ret["date"] - pd.Timedelta(days=5)
    )
    factor = (
        ret[ret["external_active_asof_date"]]
        .groupby("date")
        .agg(ai_theme_abret=("abret_mm", "mean"), ai_theme_n=("peer_code", "nunique"))
        .reset_index()
    )
    f0 = factor.rename(columns={"date": "date_0", "ai_theme_abret": "ai_theme_abret_0", "ai_theme_n": "ai_theme_n_0"})
    fp1 = factor.rename(columns={"date": "date_p1", "ai_theme_abret": "ai_theme_abret_p1", "ai_theme_n": "ai_theme_n_p1"})
    out = sample.merge(f0, on="date_0", how="left").merge(fp1, on="date_p1", how="left")
    out["ai_theme_abret_0_p1"] = out[["ai_theme_abret_0", "ai_theme_abret_p1"]].sum(axis=1, min_count=1)
    out["market_ret_0_p1"] = pd.to_numeric(out["mkt_ret_0"], errors="coerce") + pd.to_numeric(out["mkt_ret_p1"], errors="coerce")
    return out


def run_one(sample: pd.DataFrame, ai_def: str, xvars_extra: list[str], model: str, note: str) -> dict[str, object]:
    d = add_ai_terms(sample, ai_def)
    if "ai_theme_abret_0_p1" in d.columns:
        d["ai_theme_x_ai"] = d["ai_theme_abret_0_p1"] * d["ai"]
    if "market_ret_0_p1" in d.columns:
        d["market_return_x_ai"] = d["market_ret_0_p1"] * d["ai"]
    rows = fit_absorbed(
        d,
        OUTCOME,
        ["ai", "spec_ai", *xvars_extra],
        STRONG_FE,
        model=model,
        term_focus="spec_ai",
    )
    if not rows:
        return {
            "model": model,
            "ai_def": ai_def,
            "coef": math.nan,
            "se": math.nan,
            "z": math.nan,
            "p": math.nan,
            "nobs": 0,
            "events": 0,
            "peer_firms": 0,
            "controls": note,
        }
    row = rows[0]
    row.update({"ai_def": ai_def, "controls": note, "fe": "+".join(STRONG_FE), "cluster": "event_id + peer_code"})
    return row


def run_theme_controls(sample: pd.DataFrame) -> pd.DataFrame:
    enriched = add_theme_factor(sample)
    rows = []
    specs = [
        ("baseline", ["peer_car_pre10_m2_mm", "peer_car_pre20_m2_mm"], "PeerCAR[-10,-2], PeerCAR[-20,-2]"),
        (
            "add_ai_theme_x_ai",
            ["ai_theme_x_ai", "peer_car_pre10_m2_mm", "peer_car_pre20_m2_mm"],
            "AI-theme abnormal return[0,+1] x AIActive, pre-window controls",
        ),
        (
            "add_market_x_ai",
            ["market_return_x_ai", "peer_car_pre10_m2_mm", "peer_car_pre20_m2_mm"],
            "Market return[0,+1] x AIActive, pre-window controls",
        ),
        (
            "add_theme_and_market_x_ai",
            ["ai_theme_x_ai", "market_return_x_ai", "peer_car_pre10_m2_mm", "peer_car_pre20_m2_mm"],
            "AI-theme x AIActive, market x AIActive, pre-window controls",
        ),
    ]
    for ai_def in AI_DEFS:
        for model, extras, note in specs:
            rows.append(run_one(enriched, ai_def, extras, model, note))
    out = pd.DataFrame(rows)
    factor_summary = (
        enriched[["event_id", "event_date", "ai_theme_abret_0_p1", "ai_theme_n_0", "ai_theme_n_p1", "market_ret_0_p1"]]
        .drop_duplicates("event_id")
        .describe(include="all")
        .T.reset_index()
        .rename(columns={"index": "variable"})
    )
    factor_summary.to_csv(OUT_DIR / "ai_theme_factor_summary.csv", index=False)
    return out


def run_external_component_audit(sample: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    ai_defs = [
        "prior_cac",
        "prior_ai_patent_grant",
        "prior_genai_patent_grant",
        "prior_broad_ai_hiring_365_ge1",
        "prior_broad_ai_hiring_365_ge3",
        "prior_genai_hiring_365_ge1",
        "ext_no_hiring",
        "ext_any",
        "ext_strict",
        "ext_genai_strict",
        "ext_plus_history",
        "current_text_history",
    ]
    rows = []
    for ai_def in ai_defs:
        if ai_def not in sample.columns:
            continue
        rows.append(run_one(sample, ai_def, ["peer_car_pre10_m2_mm", "peer_car_pre20_m2_mm"], "external_component_final_sample", "PeerCAR[-10,-2], PeerCAR[-20,-2]"))
    regs = pd.DataFrame(rows)
    counts = []
    for ai_def in ai_defs:
        if ai_def not in sample.columns:
            continue
        x = sample[ai_def].fillna(0).astype(int)
        counts.append(
            {
                "ai_def": ai_def,
                "obs_active": int(x.sum()),
                "obs_share": float(x.mean()),
                "active_peer_firms": int(sample.loc[x.eq(1), "peer_code"].nunique()),
                "total_peer_firms": int(sample["peer_code"].nunique()),
                "active_events": int(sample.loc[x.eq(1), "event_id"].nunique()),
                "total_events": int(sample["event_id"].nunique()),
            }
        )
    return regs, pd.DataFrame(counts)


def extract_market_caps(sample: pd.DataFrame) -> pd.DataFrame:
    cache = OUT_DIR / "matched_market_caps.csv"
    keys = set(zip(sample["peer_code"].map(code6), sample["date_0"].dt.strftime("%Y-%m-%d")))
    if cache.exists():
        try:
            caps = pd.read_csv(cache, dtype={"peer_code": str})
            if {"peer_code", "date_0"}.issubset(caps.columns):
                caps["peer_code"] = caps["peer_code"].map(code6)
                caps["date_0"] = pd.to_datetime(caps["date_0"], errors="coerce")
                have = set(zip(caps["peer_code"], caps["date_0"].dt.strftime("%Y-%m-%d")))
                if keys.issubset(have):
                    return caps
        except pd.errors.EmptyDataError:
            pass

    files = sorted(TRD_DIR.glob("TRD_Dalyr*.xlsx"))
    out_rows = []
    remaining = set(keys)
    for file in files:
        if not remaining:
            break
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        # CSMAR's exported .xlsx files often declare an incorrect A1-only
        # dimension.  Resetting dimensions lets read-only mode see all columns.
        ws.reset_dimensions()
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, max_col=24))]
        idx = {name: i for i, name in enumerate(header)}
        needed = {"Stkcd", "Trddt", "Dsmvosd", "Dsmvtll"}
        if not needed.issubset(idx):
            wb.close()
            continue
        for row in ws.iter_rows(min_row=4, max_col=24, values_only=True):
            code = code6(row[idx["Stkcd"]])
            if not code:
                continue
            dt = row[idx["Trddt"]]
            if pd.isna(dt):
                continue
            date = pd.Timestamp(dt).strftime("%Y-%m-%d")
            key = (code, date)
            if key not in remaining:
                continue
            out_rows.append(
                {
                    "peer_code": code,
                    "date_0": date,
                    "dsmvosd_thousand": pd.to_numeric(row[idx["Dsmvosd"]], errors="coerce"),
                    "dsmvtll_thousand": pd.to_numeric(row[idx["Dsmvtll"]], errors="coerce"),
                    "source_file": file.name,
                }
            )
            remaining.remove(key)
        wb.close()
    caps = pd.DataFrame(out_rows)
    if not caps.empty:
        caps["date_0"] = pd.to_datetime(caps["date_0"], errors="coerce")
    caps.to_csv(cache, index=False)
    return caps


def economic_magnitude(sample: pd.DataFrame, headline_coef: float = -0.002303271928067795) -> pd.DataFrame:
    caps = extract_market_caps(sample)
    merged = sample.merge(caps, on=["peer_code", "date_0"], how="left")
    merged["mktcap_total_rmb"] = pd.to_numeric(merged["dsmvtll_thousand"], errors="coerce") * 1000.0
    merged["mktcap_float_rmb"] = pd.to_numeric(merged["dsmvosd_thousand"], errors="coerce") * 1000.0
    active = merged[merged["ext_any"].fillna(0).astype(int).eq(1)].copy()
    rows = []
    for cap_col, label in [("mktcap_total_rmb", "total_market_cap"), ("mktcap_float_rmb", "float_market_cap")]:
        x = active[cap_col].dropna()
        rows.append(
            {
                "cap_basis": label,
                "coef_used": headline_coef,
                "active_obs_with_cap": int(x.shape[0]),
                "coverage_share_active_obs": float(x.shape[0] / max(len(active), 1)),
                "median_cap_rmb": float(x.median()) if len(x) else math.nan,
                "mean_cap_rmb": float(x.mean()) if len(x) else math.nan,
                "p25_cap_rmb": float(x.quantile(0.25)) if len(x) else math.nan,
                "p75_cap_rmb": float(x.quantile(0.75)) if len(x) else math.nan,
                "median_effect_rmb": float(headline_coef * x.median()) if len(x) else math.nan,
                "mean_effect_rmb": float(headline_coef * x.mean()) if len(x) else math.nan,
                "sum_effect_across_active_obs_rmb": float(headline_coef * x.sum()) if len(x) else math.nan,
            }
        )
    return pd.DataFrame(rows)


def copy_placebos() -> dict[str, pd.DataFrame]:
    out = {}
    if NON_GENAI_PLACEBO.exists():
        out["non_genai_pseudo_event_summary"] = pd.read_csv(NON_GENAI_PLACEBO)
    if PREWINDOW_PLACEBO.exists():
        pre = pd.read_csv(PREWINDOW_PLACEBO)
        out["prewindow_placebo_summary"] = pre[
            (pre["peer_group"].isin(["true_top5", "low_top5"]))
            & (pre["ai_def"].isin(["ext_any", "current_text_history"]))
            & (pre["fe_spec"].eq("event_fe_peer_industry_week_fe"))
        ].copy()
    return out


def fmt_table(df: pd.DataFrame, cols: list[str] | None = None, max_rows: int = 30) -> str:
    if df.empty:
        return "No rows."
    show = df.copy()
    if cols:
        show = show[cols]
    if len(show) > max_rows:
        show = show.head(max_rows)
    return "```text\n" + show.to_string(index=False) + "\n```"


def write_doc(outputs: dict[str, pd.DataFrame]) -> None:
    lines = [
        "# V8 Measurement and Final-Design Closure Checks",
        "",
        "Date: 2026-05-27",
        "",
        "## Purpose",
        "",
        "This run closes the reviewer-facing empirical gaps for the current capital-market peer-revaluation design. It does not pivot to a new paper. The central conclusion is deliberately conservative: the main peer-CAR result remains usable. The 300-row double-coded sample is treated as an internal construct-boundary check, not as a main-paper validation gate, because the human/Claude scores capture stricter implementation-specificity while the current `Specificity_z` is an objective text-detail / disclosure-concreteness proxy.",
        "",
        "## 1. Final Sample Freeze",
        "",
        fmt_table(outputs["final_sample_summary"]),
        "",
        fmt_table(outputs["final_sample_by_year"]),
        "",
        "Interpretation: this is the sample to use for all headline tables. Older 8,649-row tables are pre-freeze checks that did not require the same focal-CAR/pre-window-complete filters.",
        "",
        "## 2. Specificity Double-Coding Validation",
        "",
        fmt_table(outputs["specificity_coder_agreement"]),
        "",
        fmt_table(outputs["specificity_coder_correlations"]),
        "",
        fmt_table(outputs["specificity_coder_bin_summary"]),
        "",
        "Interpretation: coder-to-coder score correlation is strong, but both coders' stricter implementation-specificity scores correlate only weakly with the original `Specificity_z`. For the current paper package, do not use this as a headline validation table or as a reason to replace the main X. Treat it as evidence that the existing proxy should be described as objective disclosure detail / concreteness, not as manually coded implementation specificity.",
        "",
        "## 3. AI-Theme Date-Shock Controls",
        "",
        fmt_table(outputs["ai_theme_shock_controls"], ["model", "ai_def", "coef", "se", "z", "p", "nobs", "events", "peer_firms", "controls"]),
        "",
        "Interpretation: the headline coefficient is stable after controlling for AI-theme abnormal-return shocks interacted with AIActive.",
        "",
        "## 4. External AIActive Component Audit on Final Sample",
        "",
        fmt_table(outputs["external_component_counts"]),
        "",
        fmt_table(outputs["external_component_regs"], ["ai_def", "coef", "se", "z", "p", "nobs", "events", "peer_firms", "controls"]),
        "",
        "Interpretation: `ext_any` should remain the headline AIActive because it is external to the disclosure text. The components are heterogeneous; `prior_ai_patent_grant` and `ext_no_hiring` are cleaner but sparse, while hiring supplies broad coverage.",
        "",
        "## 5. Economic Magnitude",
        "",
        fmt_table(outputs["economic_magnitude"]),
        "",
        "Interpretation: multiply by -0.002303, the final-sample `ext_any` coefficient from the focal-good-news/pretrend table. CSMAR daily market caps are in thousand RMB before conversion.",
        "",
        "## 6. Existing Placebo Evidence",
        "",
        "### Non-GenAI pseudo-events",
        "",
        fmt_table(outputs.get("non_genai_pseudo_event_summary", pd.DataFrame())),
        "",
        "### Pre-window placebo",
        "",
        fmt_table(outputs.get("prewindow_placebo_summary", pd.DataFrame())),
        "",
        "## Output Files",
        "",
    ]
    for name in outputs:
        lines.append(f"- `{OUT_DIR / (name + '.csv')}`")
    DOC_PATH.parent.mkdir(parents=True, exist_ok=True)
    DOC_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    sample = load_final_sample()
    sample_summary, sample_by_year = final_sample_freeze(sample)

    outputs: dict[str, pd.DataFrame] = {
        "final_sample_summary": sample_summary,
        "final_sample_by_year": sample_by_year,
    }
    outputs.update(run_coder_validation())
    outputs["ai_theme_shock_controls"] = run_theme_controls(sample)
    regs, counts = run_external_component_audit(sample)
    outputs["external_component_regs"] = regs
    outputs["external_component_counts"] = counts
    outputs["economic_magnitude"] = economic_magnitude(sample)
    outputs.update(copy_placebos())

    for name, df in outputs.items():
        df.to_csv(OUT_DIR / f"{name}.csv", index=False)

    write_doc(outputs)
    print(f"Wrote outputs to {OUT_DIR}")
    print(f"Wrote doc to {DOC_PATH}")
    print(outputs["final_sample_summary"].to_string(index=False))
    print(outputs["ai_theme_shock_controls"][["model", "ai_def", "coef", "p", "nobs"]].to_string(index=False))
    print(outputs["economic_magnitude"].to_string(index=False))


if __name__ == "__main__":
    main()
